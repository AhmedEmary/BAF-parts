import os

from odoo import http, _
from odoo.http import request
from odoo.exceptions import AccessError, MissingError
from odoo.addons.portal.controllers.portal import CustomerPortal
from odoo.tools import file_open
from io import BytesIO

from odoo.addons.general_system_custom.models.baf_product_pricing import (
    baf_brand_base_key,
    baf_family_base_key,
)

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    openpyxl = None
    Font = None


class BafDiscountTemplateDownload(http.Controller):
    """Serves the discount-matrix import template for a single brand format."""

    @http.route(
        '/general_system_custom/discount_matrix_template',
        type='http', auth='user',
    )
    def download_discount_matrix_template(self, family_id=None, method=None, **kw):
        if not openpyxl:
            return request.not_found()
        try:
            fid = int(family_id)
        except (TypeError, ValueError):
            return request.not_found()

        family = request.env['baf.brand.family'].browse(fid).exists()
        if not family:
            return request.not_found()
        # Types & Groups: one base per brand (each split into T12/T39).
        # Groups Table: the whole family shares one base (its normalized name).
        type_split = method != 'groups_only'
        if type_split:
            bases = []
            for brand in family.brand_ids:
                base = baf_brand_base_key(brand.name)
                if base and base not in bases:
                    bases.append(base)
        else:
            base = baf_family_base_key(family)
            bases = [base] if base else []
        if not bases:
            return request.not_found()

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        self._build_matrix_sheet(wb, bases, type_split)
        filename = 'discount_matrix_%s_template.xlsx' % '_'.join(bases)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        data = output.read()
        output.close()

        return request.make_response(
            data,
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', 'attachment; filename="%s"' % filename),
            ],
        )

    # ── Sheet builder ───────────────────────────────────────────────────
    # The layout mirrors what baf.discount.import.wizard detects, so a file
    # produced from this template can be re-uploaded without editing.

    @staticmethod
    def _bold(ws, row, col, value):
        ws.cell(row=row, column=col, value=value).font = Font(bold=True)

    def _build_matrix_sheet(self, wb, bases, with_types):
        """BMW/MINI: row 1 names each group, row 2 names the group's type
        sub-columns (one pair per brand). Single brand: row 1 names each group,
        one column each."""
        ws = wb.create_sheet('-'.join(bases))
        group_labels = ['SALE PRICE GR1', 'SALE PRICE GR2',
                        'SALE PRICE GR3', 'SALE PRICE GR4']
        self._bold(ws, 1, 1, 'DC')

        if not with_types:
            for offset, label in enumerate(group_labels):
                self._bold(ws, 1, 2 + offset, label)
            ws['A2'] = '1A'  # sample DC — codes may be alphanumeric
            return

        # T12 column covers type codes 1, 2, 4, 6, 8; T39 covers 3, 5, 7, 9.
        subs = [f'{base} TA {codes}'
                for base in bases
                for codes in ('1-2-4-6-8', '3-5-7-9')]
        # The moto tier is a group like any other, detected from its header.
        for section, label in enumerate(group_labels + ['GR_MOTORCYCLE']):
            base_col = 2 + section * len(subs)
            self._bold(ws, 1, base_col, label)
            for i, sub in enumerate(subs):
                self._bold(ws, 2, base_col + i, sub)
        ws['A3'] = '10'  # sample DC


class PortalExcelExport(CustomerPortal):

    def _generate_excel_response(self, filename, workbook):
        """Helper to create the HTTP response that downloads the file"""
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        xls_data = output.read()
        output.close()

        return request.make_response(
            xls_data,
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}.xlsx"'),
            ]
        )

    @http.route(['/my/orders/<int:order_id>/export_excel'], type='http', auth="public", website=True)
    def portal_order_export_excel(self, order_id, access_token=None, **kw):
        """Handles Excel Export from the Sales Order Website Portal"""
        try:
            order_sudo = self._document_check_access('sale.order', order_id, access_token=access_token)
        except (AccessError, MissingError):
            return request.redirect('/my')

        if not openpyxl:
            return request.redirect('/my')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Order"
        bold_font = Font(bold=True)

        headers = ["SKU", "Product", "Quantity", "Unit Price", "Subtotal"]
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_num).font = bold_font

        for line in order_sudo.order_line:
            ws.append([
                line.product_id.default_code or "",
                line.product_id.name or "",
                line.product_uom_qty,
                line.price_unit,
                line.price_subtotal
            ])

        ws.append(["", "", "", "Untaxed Amount:", order_sudo.amount_untaxed])
        ws.append(["", "", "", "Taxes:", order_sudo.amount_tax])
        ws.append(["", "", "", "Total:", order_sudo.amount_total])

        for row_idx in range(ws.max_row - 2, ws.max_row + 1):
            ws.cell(row=row_idx, column=4).font = bold_font
            ws.cell(row=row_idx, column=5).font = bold_font

        return self._generate_excel_response(f"Order_{order_sudo.name}", wb)

    @http.route(['/my/invoices/<int:invoice_id>/export_excel'], type='http', auth="public", website=True)
    def portal_invoice_export_excel(self, invoice_id, access_token=None, **kw):
        """Handles Excel Export from the Invoice Website Portal"""
        try:
            invoice_sudo = self._document_check_access('account.move', invoice_id, access_token=access_token)
        except (AccessError, MissingError):
            return request.redirect('/my')

        if not openpyxl:
            return request.redirect('/my')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"
        bold_font = Font(bold=True)

        headers = ["SKU", "Product", "Quantity", "Unit Price", "Subtotal"]
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_num).font = bold_font

        for line in invoice_sudo.invoice_line_ids.filtered(lambda l: l.display_type == 'product'):
            ws.append([
                line.product_id.default_code or "",
                line.product_id.name or "",
                line.quantity,
                line.price_unit,
                line.price_subtotal
            ])

        ws.append(["", "", "", "Untaxed Amount:", invoice_sudo.amount_untaxed])
        ws.append(["", "", "", "Taxes:", invoice_sudo.amount_tax])
        ws.append(["", "", "", "Total:", invoice_sudo.amount_total])

        for row_idx in range(ws.max_row - 2, ws.max_row + 1):
            ws.cell(row=row_idx, column=4).font = bold_font
            ws.cell(row=row_idx, column=5).font = bold_font

        return self._generate_excel_response(f"Invoice_{invoice_sudo.name.replace('/', '_')}", wb)
