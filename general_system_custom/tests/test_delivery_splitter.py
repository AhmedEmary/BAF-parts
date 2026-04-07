"""
Comprehensive tests for the Inbound Reconciliation / Delivery Splitter feature.

Models covered:
  - delivery.list.import       (parent)
  - delivery.list.line         (split lines)
  - delivery.list.excel.line   (raw Excel data)
  - purchase.order.line        (qty_open / qty_split side-effects)

Run with:
  ./odoo-bin -c odoo.conf --test-enable --stop-after-init -m general_system_custom \
      --test-tags /general_system_custom:TestDeliveryListSplitter
"""
import base64
import io
from datetime import timedelta

from odoo import Command
from odoo.exceptions import UserError
from odoo.tests import TransactionCase, tagged

try:
    import openpyxl
except ImportError:
    openpyxl = None


@tagged('post_install', '-at_install')
class TestDeliveryListSplitter(TransactionCase):
    """Full-coverage tests for the delivery splitter (Inbound Reconciliation)."""

    # ──────────────────────────────────────────────────────────────────────────
    # Class-level setup — only the supplier; products/POs are per-test for
    # isolation (avoids date-order / qty_open cross-contamination).
    # ──────────────────────────────────────────────────────────────────────────

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.vendor = cls.env['res.partner'].create({
            'name': 'Test Splitter Vendor',
            'supplier_rank': 1,
        })

    # ──────────────────────────────────────────────────────────────────────────
    # Helpers
    # ──────────────────────────────────────────────────────────────────────────

    def _make_excel(self, rows):
        """Return base64-encoded xlsx bytes.

        rows: list of (sku, qty, price) tuples — header is added automatically.
        """
        if not openpyxl:
            self.skipTest('openpyxl not installed')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['SKU', 'Qty', 'Price'])
        for row in rows:
            ws.append(list(row))
        buf = io.BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue()).decode()

    def _make_import(self, rows=None, supplier=None):
        """Create a DeliveryListImport, optionally with an attached Excel file."""
        vals = {'supplier_id': (supplier or self.vendor).id}
        if rows is not None:
            vals['file_data'] = self._make_excel(rows)
            vals['file_name'] = 'test.xlsx'
        return self.env['delivery.list.import'].create(vals)

    def _process(self, rows, supplier=None):
        """Create + action_process_file in one step."""
        imp = self._make_import(rows, supplier=supplier)
        imp.action_process_file()
        return imp

    def _make_product(self, sku, price_unit=10.0):
        """Create a storable product with the given default_code."""
        return self.env['product.product'].create({
            'name': f'Product {sku}',
            'type': 'consu',
            'is_storable': True,
            'default_code': sku,
            'standard_price': price_unit,
        })

    def _make_po(self, product, qty, price_unit=10.0, supplier=None, date_offset_days=0):
        """Create and confirm a PO for one product line."""
        po = self.env['purchase.order'].create({
            'partner_id': (supplier or self.vendor).id,
            'order_line': [Command.create({
                'product_id': product.id,
                'product_qty': qty,
                'price_unit': price_unit,
            })],
        })
        po.button_confirm()
        if date_offset_days:
            po.write({'date_order': po.date_order + timedelta(days=date_offset_days)})
        return po

    # ──────────────────────────────────────────────────────────────────────────
    # action_process_file — happy paths
    # ──────────────────────────────────────────────────────────────────────────

    def test_process_sets_state_to_processed(self):
        """action_process_file transitions state draft → processed."""
        p = self._make_product('TST-P01')
        self._make_po(p, 10.0)
        imp = self._process([('TST-P01', 5, 10.0)])
        self.assertEqual(imp.state, 'processed')

    def test_process_creates_split_lines(self):
        """One SKU + one PO → one split line created."""
        p = self._make_product('TST-P02')
        po = self._make_po(p, 10.0)
        imp = self._process([('TST-P02', 4, 10.0)])
        self.assertEqual(len(imp.line_ids), 1)
        line = imp.line_ids[0]
        self.assertEqual(line.product_id, p)
        self.assertEqual(line.po_id, po)
        self.assertAlmostEqual(line.qty_split, 4.0)

    def test_process_creates_excel_lines(self):
        """Raw Excel data is persisted in excel_line_ids."""
        p1 = self._make_product('TST-P03A')
        p2 = self._make_product('TST-P03B')
        po = self.env['purchase.order'].create({
            'partner_id': self.vendor.id,
            'order_line': [
                Command.create({'product_id': p1.id, 'product_qty': 5.0, 'price_unit': 1.0}),
                Command.create({'product_id': p2.id, 'product_qty': 5.0, 'price_unit': 1.0}),
            ],
        })
        po.button_confirm()
        imp = self._process([('TST-P03A', 3, 1.0), ('TST-P03B', 2, 1.0)])
        self.assertEqual(len(imp.excel_line_ids), 2)
        skus = set(imp.excel_line_ids.mapped('sku'))
        self.assertIn('TST-P03A', skus)
        self.assertIn('TST-P03B', skus)

    def test_process_fifo_splits_across_pos(self):
        """FIFO: 7 units delivered, PO1=5 (older) + PO2=8 (newer) → 5 + 2."""
        p = self._make_product('TST-P04')
        po1 = self._make_po(p, 5.0, date_offset_days=-10)  # older
        po2 = self._make_po(p, 8.0)                         # newer
        imp = self._process([('TST-P04', 7, 10.0)])

        lines = imp.line_ids.sorted('po_date')
        self.assertEqual(len(lines), 2)
        self.assertEqual(lines[0].po_id, po1)
        self.assertAlmostEqual(lines[0].qty_split, 5.0)
        self.assertEqual(lines[1].po_id, po2)
        self.assertAlmostEqual(lines[1].qty_split, 2.0)

    def test_process_overflow_appended_to_last_po(self):
        """When delivery exceeds total open qty, overflow goes to last PO line."""
        p = self._make_product('TST-P05')
        po = self._make_po(p, 5.0)
        imp = self._process([('TST-P05', 9, 10.0)])  # 9 > 5 open
        self.assertEqual(len(imp.line_ids), 1)
        self.assertAlmostEqual(imp.line_ids[0].qty_split, 9.0)

    def test_process_aggregates_duplicate_skus(self):
        """Two Excel rows with the same SKU are aggregated into one excel line."""
        p = self._make_product('TST-P06')
        self._make_po(p, 20.0)
        # Build Excel with 2 rows for same SKU
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['SKU', 'Qty', 'Price'])
        ws.append(['TST-P06', 3, 10.0])
        ws.append(['TST-P06', 4, 10.0])  # duplicate row
        buf = io.BytesIO()
        wb.save(buf)
        imp = self._make_import()
        imp.write({
            'file_data': base64.b64encode(buf.getvalue()).decode(),
            'file_name': 'test.xlsx',
        })
        imp.action_process_file()
        self.assertEqual(len(imp.excel_line_ids), 1)
        self.assertAlmostEqual(imp.excel_line_ids[0].qty_received, 7.0)
        self.assertAlmostEqual(imp.line_ids[0].qty_split, 7.0)

    def test_process_skips_zero_qty_rows(self):
        """Rows with qty=0 are silently ignored."""
        p = self._make_product('TST-P07')
        self._make_po(p, 10.0)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['SKU', 'Qty', 'Price'])
        ws.append(['TST-P07', 5, 10.0])
        ws.append(['TST-P07', 0, 10.0])  # zero qty — must be skipped
        buf = io.BytesIO()
        wb.save(buf)
        imp = self._make_import()
        imp.write({
            'file_data': base64.b64encode(buf.getvalue()).decode(),
            'file_name': 'test.xlsx',
        })
        imp.action_process_file()
        self.assertAlmostEqual(imp.excel_line_ids[0].qty_received, 5.0)

    def test_process_clears_existing_lines_on_reprocess(self):
        """After reset-to-draft + re-upload, old lines are cleared and replaced."""
        p = self._make_product('TST-P08')
        self._make_po(p, 20.0)
        imp = self._make_import([('TST-P08', 5, 10.0)])
        imp.action_process_file()
        old_line_id = imp.line_ids[0].id

        imp.action_reset_to_draft()
        imp.write({'file_data': self._make_excel([('TST-P08', 8, 11.0)])})
        imp.action_process_file()

        self.assertFalse(
            self.env['delivery.list.line'].browse(old_line_id).exists(),
            'Old split line should have been deleted',
        )
        self.assertEqual(len(imp.line_ids), 1)
        self.assertAlmostEqual(imp.line_ids[0].qty_split, 8.0)

    def test_process_excel_sequence_preserved(self):
        """excel_sequence on split lines reflects the SKU's first appearance order."""
        p1 = self._make_product('TST-P09A')
        p2 = self._make_product('TST-P09B')
        po = self.env['purchase.order'].create({
            'partner_id': self.vendor.id,
            'order_line': [
                Command.create({'product_id': p1.id, 'product_qty': 5.0, 'price_unit': 1.0}),
                Command.create({'product_id': p2.id, 'product_qty': 5.0, 'price_unit': 1.0}),
            ],
        })
        po.button_confirm()
        # p2 appears first in Excel, p1 second
        imp = self._process([('TST-P09B', 2, 1.0), ('TST-P09A', 3, 1.0)])
        line_a = imp.line_ids.filtered(lambda l: l.product_id == p1)
        line_b = imp.line_ids.filtered(lambda l: l.product_id == p2)
        self.assertGreater(line_a[0].excel_sequence, line_b[0].excel_sequence)

    # ──────────────────────────────────────────────────────────────────────────
    # action_process_file — error paths
    # ──────────────────────────────────────────────────────────────────────────

    def test_process_raises_no_file(self):
        """UserError when no Excel file is attached."""
        imp = self._make_import()  # no file
        with self.assertRaises(UserError):
            imp.action_process_file()

    def test_process_raises_empty_excel(self):
        """UserError when Excel contains only headers and no data rows."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['SKU', 'Qty', 'Price'])
        buf = io.BytesIO()
        wb.save(buf)
        imp = self._make_import()
        imp.write({
            'file_data': base64.b64encode(buf.getvalue()).decode(),
            'file_name': 'test.xlsx',
        })
        with self.assertRaises(UserError):
            imp.action_process_file()

    def test_process_raises_missing_columns(self):
        """UserError when required columns (SKU/Qty/Price) are absent."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Product', 'Amount'])  # wrong column names
        ws.append(['SOMETHING', 5])
        buf = io.BytesIO()
        wb.save(buf)
        imp = self._make_import()
        imp.write({
            'file_data': base64.b64encode(buf.getvalue()).decode(),
            'file_name': 'test.xlsx',
        })
        with self.assertRaises(UserError):
            imp.action_process_file()

    def test_process_raises_unknown_sku(self):
        """UserError when a SKU does not match any product in the system."""
        imp = self._make_import([('NO-SUCH-SKU-XYZ', 5, 10.0)])
        with self.assertRaises(UserError):
            imp.action_process_file()

    def test_process_raises_no_po_for_product(self):
        """UserError when product exists but supplier has no confirmed PO for it."""
        p = self._make_product('TST-ORPHAN-01')
        # No PO created for this product
        imp = self._make_import([('TST-ORPHAN-01', 5, 10.0)])
        with self.assertRaises(UserError):
            imp.action_process_file()

    def test_process_raises_all_pos_fully_allocated(self):
        """UserError when every PO line for that product has qty_open == 0."""
        p = self._make_product('TST-FULL-01')
        po = self._make_po(p, 5.0)
        po.order_line[0].write({'qty_split': 5.0})  # fully allocated
        imp = self._make_import([('TST-FULL-01', 3, 10.0)])
        with self.assertRaises(UserError):
            imp.action_process_file()

    # ──────────────────────────────────────────────────────────────────────────
    # action_confirm_delivery
    # ──────────────────────────────────────────────────────────────────────────

    def test_confirm_updates_po_line_qty_split(self):
        """Confirming adds qty_split to the PO line."""
        p = self._make_product('TST-C01')
        po = self._make_po(p, 10.0)
        po_line = po.order_line[0]
        self.assertAlmostEqual(po_line.qty_split, 0.0)

        imp = self._process([('TST-C01', 4, 10.0)])
        imp.action_confirm_delivery()

        po_line.invalidate_recordset()
        self.assertAlmostEqual(po_line.qty_split, 4.0)

    def test_confirm_reduces_qty_open(self):
        """After confirmation, po_line.qty_open = ordered - split."""
        p = self._make_product('TST-C02')
        po = self._make_po(p, 10.0)
        po_line = po.order_line[0]
        self.assertAlmostEqual(po_line.qty_open, 10.0)

        imp = self._process([('TST-C02', 6, 10.0)])
        imp.action_confirm_delivery()

        po_line.invalidate_recordset()
        self.assertAlmostEqual(po_line.qty_open, 4.0)

    def test_confirm_sets_state_confirmed(self):
        p = self._make_product('TST-C03')
        self._make_po(p, 10.0)
        imp = self._process([('TST-C03', 5, 10.0)])
        imp.action_confirm_delivery()
        self.assertEqual(imp.state, 'confirmed')

    def test_confirm_returns_client_action(self):
        """action_confirm_delivery returns an ir.actions.client notification dict."""
        p = self._make_product('TST-C04')
        self._make_po(p, 10.0)
        imp = self._process([('TST-C04', 5, 10.0)])
        result = imp.action_confirm_delivery()
        self.assertEqual(result.get('type'), 'ir.actions.client')
        self.assertEqual(result.get('tag'), 'display_notification')
        # next action must include res_id so the UI can reload the form
        next_action = result['params']['next']
        self.assertEqual(next_action.get('res_model'), 'delivery.list.import')
        self.assertEqual(next_action.get('res_id'), imp.id)

    def test_confirm_raises_no_lines(self):
        """UserError when trying to confirm with no split lines."""
        imp = self._make_import()
        imp.write({'state': 'processed'})
        with self.assertRaises(UserError):
            imp.action_confirm_delivery()

    def test_confirm_raises_qty_mismatch(self):
        """UserError when sum(qty_split) != qty_received from Excel."""
        p = self._make_product('TST-C05')
        self._make_po(p, 10.0)
        imp = self._process([('TST-C05', 5, 10.0)])
        # Corrupt the split qty
        imp.line_ids[0].write({'qty_split': 3.0})  # 3 != 5
        with self.assertRaises(UserError):
            imp.action_confirm_delivery()

    def test_confirm_groups_lines_on_same_po_line(self):
        """Multiple split lines pointing to the same PO line are summed correctly."""
        p = self._make_product('TST-C06')
        po = self._make_po(p, 20.0)
        po_line = po.order_line[0]

        imp = self._process([('TST-C06', 6, 10.0)])
        # Manually add a second line targeting the same po_line and adjust excel qty
        self.env['delivery.list.line'].create({
            'import_id': imp.id,
            'product_id': p.id,
            'qty_split': 4.0,
            'po_line_id': po_line.id,
            'po_id': po.id,
            'price_supplier': 10.0,
            'price_po': 10.0,
        })
        imp.excel_line_ids[0].write({'qty_received': 10.0})  # 6 + 4 = 10

        imp.action_confirm_delivery()
        po_line.invalidate_recordset()
        self.assertAlmostEqual(po_line.qty_split, 10.0)

    def test_confirm_across_multiple_po_lines(self):
        """FIFO split across 2 POs: both PO lines get their qty_split updated."""
        p = self._make_product('TST-C07')
        po1 = self._make_po(p, 5.0, date_offset_days=-10)
        po2 = self._make_po(p, 8.0)
        po1_line = po1.order_line[0]
        po2_line = po2.order_line[0]

        imp = self._process([('TST-C07', 7, 10.0)])
        imp.action_confirm_delivery()

        po1_line.invalidate_recordset()
        po2_line.invalidate_recordset()
        self.assertAlmostEqual(po1_line.qty_split, 5.0)
        self.assertAlmostEqual(po2_line.qty_split, 2.0)

    # ──────────────────────────────────────────────────────────────────────────
    # action_reset_to_draft
    # ──────────────────────────────────────────────────────────────────────────

    def test_reset_to_draft_clears_split_lines(self):
        p = self._make_product('TST-R01')
        self._make_po(p, 10.0)
        imp = self._process([('TST-R01', 5, 10.0)])
        self.assertTrue(imp.line_ids)
        imp.action_reset_to_draft()
        self.assertFalse(imp.line_ids)

    def test_reset_to_draft_clears_excel_lines(self):
        p = self._make_product('TST-R02')
        self._make_po(p, 10.0)
        imp = self._process([('TST-R02', 5, 10.0)])
        self.assertTrue(imp.excel_line_ids)
        imp.action_reset_to_draft()
        self.assertFalse(imp.excel_line_ids)

    def test_reset_to_draft_sets_state_draft(self):
        p = self._make_product('TST-R03')
        self._make_po(p, 10.0)
        imp = self._process([('TST-R03', 5, 10.0)])
        imp.action_reset_to_draft()
        self.assertEqual(imp.state, 'draft')

    def test_reset_to_draft_preserves_file_data(self):
        """File attachment is kept after reset — user can re-process or replace."""
        p = self._make_product('TST-R04')
        self._make_po(p, 10.0)
        imp = self._make_import([('TST-R04', 5, 10.0)])
        file_before = imp.file_data
        imp.action_process_file()
        imp.action_reset_to_draft()
        self.assertEqual(imp.file_data, file_before)

    def test_reset_to_draft_raises_if_confirmed(self):
        """Can only reset-to-draft from 'processed'; confirmed state is blocked."""
        p = self._make_product('TST-R05')
        self._make_po(p, 10.0)
        imp = self._process([('TST-R05', 5, 10.0)])
        imp.action_confirm_delivery()
        with self.assertRaises(UserError):
            imp.action_reset_to_draft()

    def test_reset_to_draft_raises_if_already_draft(self):
        """Cannot reset-to-draft when already in draft state."""
        imp = self._make_import()
        with self.assertRaises(UserError):
            imp.action_reset_to_draft()

    # ──────────────────────────────────────────────────────────────────────────
    # action_reset_to_processed
    # ──────────────────────────────────────────────────────────────────────────

    def test_reset_to_processed_subtracts_qty_split(self):
        """Reset-to-processed reverses the qty_split added by confirmation."""
        p = self._make_product('TST-RP01')
        po = self._make_po(p, 10.0)
        po_line = po.order_line[0]

        imp = self._process([('TST-RP01', 6, 10.0)])
        imp.action_confirm_delivery()
        po_line.invalidate_recordset()
        self.assertAlmostEqual(po_line.qty_split, 6.0)

        imp.action_reset_to_processed()
        po_line.invalidate_recordset()
        self.assertAlmostEqual(po_line.qty_split, 0.0)

    def test_reset_to_processed_restores_qty_open(self):
        """After reset-to-processed, qty_open returns to its pre-confirm value."""
        p = self._make_product('TST-RP02')
        po = self._make_po(p, 10.0)
        po_line = po.order_line[0]
        open_before = po_line.qty_open  # 10

        imp = self._process([('TST-RP02', 6, 10.0)])
        imp.action_confirm_delivery()
        imp.action_reset_to_processed()

        po_line.invalidate_recordset()
        self.assertAlmostEqual(po_line.qty_open, open_before)

    def test_reset_to_processed_sets_state_processed(self):
        p = self._make_product('TST-RP03')
        self._make_po(p, 10.0)
        imp = self._process([('TST-RP03', 5, 10.0)])
        imp.action_confirm_delivery()
        imp.action_reset_to_processed()
        self.assertEqual(imp.state, 'processed')

    def test_reset_to_processed_raises_if_not_confirmed(self):
        """Can only reset-to-processed from 'confirmed'; processed state is blocked."""
        p = self._make_product('TST-RP04')
        self._make_po(p, 10.0)
        imp = self._process([('TST-RP04', 5, 10.0)])  # state = processed
        with self.assertRaises(UserError):
            imp.action_reset_to_processed()

    def test_reset_to_processed_raises_if_negative_would_result(self):
        """Safety check: blocked if po_line.qty_split would go negative."""
        p = self._make_product('TST-RP05')
        po = self._make_po(p, 10.0)
        imp = self._process([('TST-RP05', 5, 10.0)])
        imp.action_confirm_delivery()
        # Externally zero out qty_split (simulate it was used by another process)
        po.order_line[0].write({'qty_split': 0.0})
        with self.assertRaises(UserError):
            imp.action_reset_to_processed()

    def test_reset_to_processed_returns_client_action(self):
        p = self._make_product('TST-RP06')
        self._make_po(p, 10.0)
        imp = self._process([('TST-RP06', 5, 10.0)])
        imp.action_confirm_delivery()
        result = imp.action_reset_to_processed()
        self.assertEqual(result.get('type'), 'ir.actions.client')

    # ──────────────────────────────────────────────────────────────────────────
    # Compute — allocation flags
    # ──────────────────────────────────────────────────────────────────────────

    def test_allocation_flag_exact_fit(self):
        """qty_split == open_qty_po → neither flag set."""
        p = self._make_product('TST-AF01')
        self._make_po(p, 5.0)
        imp = self._process([('TST-AF01', 5, 10.0)])
        line = imp.line_ids[0]
        self.assertFalse(line.is_under_allocation)
        self.assertFalse(line.is_over_allocation)

    def test_allocation_flag_under(self):
        """qty_split < open_qty_po → is_under_allocation."""
        p = self._make_product('TST-AF02')
        self._make_po(p, 10.0)
        imp = self._process([('TST-AF02', 3, 10.0)])  # 3 < 10
        line = imp.line_ids[0]
        self.assertTrue(line.is_under_allocation)
        self.assertFalse(line.is_over_allocation)

    def test_allocation_flag_over(self):
        """qty_split > open_qty_po → is_over_allocation."""
        p = self._make_product('TST-AF03')
        self._make_po(p, 5.0)
        imp = self._process([('TST-AF03', 8, 10.0)])  # overflow: 8 > 5
        line = imp.line_ids[0]
        self.assertFalse(line.is_under_allocation)
        self.assertTrue(line.is_over_allocation)

    def test_allocation_flags_cleared_after_manual_edit(self):
        """Manually setting qty_split to exact open qty clears both flags."""
        p = self._make_product('TST-AF04')
        po = self._make_po(p, 5.0)
        imp = self._process([('TST-AF04', 3, 10.0)])  # starts under
        line = imp.line_ids[0]
        self.assertTrue(line.is_under_allocation)
        # Set to exact open qty
        line.write({'qty_split': 5.0})
        line.invalidate_recordset()
        self.assertFalse(line.is_under_allocation)
        self.assertFalse(line.is_over_allocation)

    # ──────────────────────────────────────────────────────────────────────────
    # Compute — price variance
    # ──────────────────────────────────────────────────────────────────────────

    def test_price_variance_detected(self):
        """price_supplier != price_po → price_variance=True, difference shown."""
        p = self._make_product('TST-PV01')
        self._make_po(p, 5.0, price_unit=10.0)
        imp = self._process([('TST-PV01', 5, 15.0)])  # supplier=15, PO=10
        line = imp.line_ids[0]
        self.assertTrue(line.price_variance)
        self.assertAlmostEqual(line.price_difference, 5.0)

    def test_price_variance_none(self):
        """Same prices → price_variance=False."""
        p = self._make_product('TST-PV02')
        self._make_po(p, 5.0, price_unit=10.0)
        imp = self._process([('TST-PV02', 5, 10.0)])
        line = imp.line_ids[0]
        self.assertFalse(line.price_variance)
        self.assertAlmostEqual(line.price_difference, 0.0)

    def test_price_variance_negative_difference(self):
        """Supplier charges less than PO price → negative price_difference."""
        p = self._make_product('TST-PV03')
        self._make_po(p, 5.0, price_unit=20.0)
        imp = self._process([('TST-PV03', 5, 15.0)])  # supplier=15, PO=20
        line = imp.line_ids[0]
        self.assertTrue(line.price_variance)
        self.assertAlmostEqual(line.price_difference, -5.0)

    # ──────────────────────────────────────────────────────────────────────────
    # Compute — summary on parent
    # ──────────────────────────────────────────────────────────────────────────

    def test_summary_total_received_and_po_lines(self):
        p = self._make_product('TST-S01')
        self._make_po(p, 20.0)
        imp = self._process([('TST-S01', 7, 10.0)])
        self.assertEqual(imp.total_received, 7)
        self.assertEqual(imp.total_po_lines, 1)

    def test_summary_total_received_across_two_lines(self):
        """total_received sums qty_split from all lines."""
        p = self._make_product('TST-S02')
        self._make_po(p, 20.0, date_offset_days=-10)
        self._make_po(p, 20.0)
        imp = self._process([('TST-S02', 12, 10.0)])  # splits: 20+20 open, takes 12
        self.assertEqual(imp.total_received, 12)

    def test_summary_has_price_variance(self):
        p = self._make_product('TST-S03')
        self._make_po(p, 10.0, price_unit=10.0)
        imp = self._process([('TST-S03', 5, 99.0)])
        self.assertTrue(imp.has_price_variance)

    def test_summary_no_price_variance(self):
        p = self._make_product('TST-S04')
        self._make_po(p, 10.0, price_unit=10.0)
        imp = self._process([('TST-S04', 5, 10.0)])
        self.assertFalse(imp.has_price_variance)

    def test_summary_has_under_allocation(self):
        p = self._make_product('TST-S05')
        self._make_po(p, 10.0)
        imp = self._process([('TST-S05', 3, 10.0)])  # 3 < 10 open
        self.assertTrue(imp.has_under_allocation)
        self.assertFalse(imp.has_over_allocation)

    def test_summary_has_over_allocation(self):
        p = self._make_product('TST-S06')
        self._make_po(p, 5.0)
        imp = self._process([('TST-S06', 8, 10.0)])  # overflow
        self.assertFalse(imp.has_under_allocation)
        self.assertTrue(imp.has_over_allocation)

    def test_summary_cleared_after_reset_to_draft(self):
        p = self._make_product('TST-S07')
        self._make_po(p, 10.0)
        imp = self._process([('TST-S07', 3, 10.0)])
        self.assertEqual(imp.total_po_lines, 1)
        imp.action_reset_to_draft()
        self.assertEqual(imp.total_po_lines, 0)
        self.assertEqual(imp.total_received, 0)

    # ──────────────────────────────────────────────────────────────────────────
    # Compute — sku, price_supplier, po_line_id, price_po
    # ──────────────────────────────────────────────────────────────────────────

    def test_compute_sku_matches_excel_sku(self):
        """Split line sku equals the SKU column from the Excel file."""
        p = self._make_product('TST-K01')
        self._make_po(p, 10.0)
        imp = self._process([('TST-K01', 5, 10.0)])
        self.assertEqual(imp.line_ids[0].sku, 'TST-K01')

    def test_compute_price_supplier_from_excel(self):
        """price_supplier is taken from the Price column of the Excel file."""
        p = self._make_product('TST-K02')
        self._make_po(p, 10.0)
        imp = self._process([('TST-K02', 5, 42.50)])
        self.assertAlmostEqual(imp.line_ids[0].price_supplier, 42.50)

    def test_compute_po_line_id_resolved(self):
        """po_line_id points to the correct purchase.order.line."""
        p = self._make_product('TST-K03')
        po = self._make_po(p, 10.0, price_unit=7.0)
        imp = self._process([('TST-K03', 5, 7.0)])
        self.assertEqual(imp.line_ids[0].po_line_id, po.order_line[0])

    def test_compute_price_po_from_po_line(self):
        """price_po reflects the PO line's price_unit."""
        p = self._make_product('TST-K04')
        self._make_po(p, 10.0, price_unit=33.0)
        imp = self._process([('TST-K04', 5, 33.0)])
        self.assertAlmostEqual(imp.line_ids[0].price_po, 33.0)

    def test_compute_sku_updates_when_product_changes(self):
        """Changing product_id on a split line updates sku via compute."""
        p1 = self._make_product('TST-K05A')
        p2 = self._make_product('TST-K05B')
        po = self.env['purchase.order'].create({
            'partner_id': self.vendor.id,
            'order_line': [
                Command.create({'product_id': p1.id, 'product_qty': 5.0, 'price_unit': 1.0}),
                Command.create({'product_id': p2.id, 'product_qty': 5.0, 'price_unit': 1.0}),
            ],
        })
        po.button_confirm()
        imp = self._process([('TST-K05A', 3, 1.0), ('TST-K05B', 2, 1.0)])
        line = imp.line_ids.filtered(lambda l: l.product_id == p1)[0]
        self.assertEqual(line.sku, 'TST-K05A')
        # Change product on the line
        line.write({'product_id': p2.id})
        line.invalidate_recordset()
        self.assertEqual(line.sku, 'TST-K05B')

    # ──────────────────────────────────────────────────────────────────────────
    # Compute — qty_open on purchase.order.line
    # ──────────────────────────────────────────────────────────────────────────

    def test_qty_open_formula(self):
        """qty_open = product_qty - qty_received - qty_split."""
        p = self._make_product('TST-QO01')
        po = self._make_po(p, 10.0)
        po_line = po.order_line[0]
        self.assertAlmostEqual(po_line.qty_open, 10.0)

        po_line.write({'qty_split': 3.0})
        po_line.invalidate_recordset()
        self.assertAlmostEqual(po_line.qty_open, 7.0)

    def test_open_qty_is_live_not_snapshot(self):
        """open_qty_po on the split line reflects the current po_line.qty_open (live)."""
        p = self._make_product('TST-QO02')
        po = self._make_po(p, 10.0)
        po_line = po.order_line[0]

        imp = self._process([('TST-QO02', 4, 10.0)])
        line = imp.line_ids[0]
        # Manually increase qty_split on PO (simulates another import being confirmed)
        po_line.write({'qty_split': 2.0})
        line.invalidate_recordset()
        # open_qty_po should now show 10 - 2 = 8 (live), not the original 10
        self.assertAlmostEqual(line.open_qty_po, 8.0)

    # ──────────────────────────────────────────────────────────────────────────
    # Cascading selects — available_product_ids
    # ──────────────────────────────────────────────────────────────────────────

    def test_available_products_all_excel_when_no_po_selected(self):
        """With no PO pre-selected, available_product_ids = all Excel products."""
        p1 = self._make_product('TST-CA01A')
        p2 = self._make_product('TST-CA01B')
        po = self.env['purchase.order'].create({
            'partner_id': self.vendor.id,
            'order_line': [
                Command.create({'product_id': p1.id, 'product_qty': 5.0, 'price_unit': 1.0}),
                Command.create({'product_id': p2.id, 'product_qty': 5.0, 'price_unit': 1.0}),
            ],
        })
        po.button_confirm()
        imp = self._process([('TST-CA01A', 3, 1.0), ('TST-CA01B', 2, 1.0)])
        line = imp.line_ids[0]
        # Clear po_id so compute falls back to "all Excel products"
        line.write({'po_id': False})
        line.invalidate_recordset()
        self.assertIn(p1, line.available_product_ids)
        self.assertIn(p2, line.available_product_ids)

    def test_available_products_narrowed_by_selected_po(self):
        """With a PO selected, only products on that PO appear."""
        p1 = self._make_product('TST-CA02A')
        p2 = self._make_product('TST-CA02B')
        # po_a has only p1, po_b has only p2
        po_a = self._make_po(p1, 5.0)
        po_b = self._make_po(p2, 5.0)
        imp = self._process([('TST-CA02A', 3, 1.0), ('TST-CA02B', 2, 1.0)])
        line = imp.line_ids.filtered(lambda l: l.product_id == p1)[0]
        # Switch to po_b → p1 is not on po_b
        line.write({'po_id': po_b.id})
        line.invalidate_recordset()
        self.assertNotIn(p1, line.available_product_ids)
        self.assertIn(p2, line.available_product_ids)

    def test_available_products_empty_when_no_excel_lines(self):
        """No excel_line_ids → available_product_ids is empty/falsy."""
        imp = self._make_import()  # draft, no file
        line = self.env['delivery.list.line'].create({
            'import_id': imp.id,
            'product_id': self._make_product('TST-CA03').id,
            'qty_split': 1.0,
        })
        self.assertFalse(line.available_product_ids)

    # ──────────────────────────────────────────────────────────────────────────
    # Cascading selects — available_po_ids
    # ──────────────────────────────────────────────────────────────────────────

    def test_available_po_ids_filtered_by_supplier(self):
        """available_po_ids excludes POs from a different supplier."""
        other_vendor = self.env['res.partner'].create({
            'name': 'Other Splitter Vendor',
            'supplier_rank': 1,
        })
        p = self._make_product('TST-CB01')
        po_own = self._make_po(p, 5.0)
        po_other = self._make_po(p, 5.0, supplier=other_vendor)
        imp = self._process([('TST-CB01', 3, 1.0)])
        line = imp.line_ids[0]
        self.assertIn(po_own, line.available_po_ids)
        self.assertNotIn(po_other, line.available_po_ids)

    def test_available_po_ids_only_confirmed_pos(self):
        """Draft POs are not included in available_po_ids."""
        p = self._make_product('TST-CB02')
        po_confirmed = self._make_po(p, 5.0)
        po_draft = self.env['purchase.order'].create({
            'partner_id': self.vendor.id,
            'order_line': [Command.create({
                'product_id': p.id, 'product_qty': 5.0, 'price_unit': 1.0,
            })],
        })
        # po_draft stays in draft — not confirmed
        imp = self._process([('TST-CB02', 3, 1.0)])
        line = imp.line_ids[0]
        self.assertIn(po_confirmed, line.available_po_ids)
        self.assertNotIn(po_draft, line.available_po_ids)

    def test_available_po_ids_narrowed_by_product(self):
        """When product is set, only POs containing that product appear."""
        p1 = self._make_product('TST-CB03A')
        p2 = self._make_product('TST-CB03B')
        po1 = self._make_po(p1, 5.0)
        po2 = self._make_po(p2, 5.0)
        imp = self._process([('TST-CB03A', 3, 1.0), ('TST-CB03B', 2, 1.0)])
        line = imp.line_ids.filtered(lambda l: l.product_id == p1)[0]
        self.assertIn(po1, line.available_po_ids)
        self.assertNotIn(po2, line.available_po_ids)

    def test_available_po_ids_shows_multiple_pos_for_same_product(self):
        """When a product has multiple open POs, all of them appear in available_po_ids."""
        p = self._make_product('TST-CB04')
        po1 = self._make_po(p, 5.0, date_offset_days=-5)  # older
        po2 = self._make_po(p, 5.0)                        # newer
        # Deliver 3 — FIFO assigns all to po1 (oldest)
        imp = self._process([('TST-CB04', 3, 1.0)])
        line = imp.line_ids[0]
        # Both POs must be visible in the dropdown so the user can re-assign
        self.assertIn(po1, line.available_po_ids)
        self.assertIn(po2, line.available_po_ids)

    def test_available_po_ids_empty_when_supplier_has_no_matching_pos(self):
        """A vendor with no confirmed POs for the Excel products → empty available_po_ids."""
        empty_vendor = self.env['res.partner'].create({
            'name': 'Vendor With No POs',
            'supplier_rank': 1,
        })
        p = self._make_product('TST-CB05')
        # Create a PO for p under the main vendor (not empty_vendor)
        self._make_po(p, 5.0)
        # Import under empty_vendor — it has no POs at all
        imp = self._make_import([('TST-CB05', 3, 1.0)], supplier=empty_vendor)
        # Process would fail (no PO), so manually create an excel line and a split line
        self.env['delivery.list.excel.line'].create({
            'import_id': imp.id,
            'product_id': p.id,
            'sku': 'TST-CB05',
            'qty_received': 3.0,
        })
        line = self.env['delivery.list.line'].create({
            'import_id': imp.id,
            'product_id': p.id,
            'qty_split': 3.0,
        })
        line.invalidate_recordset()
        self.assertFalse(line.available_po_ids)

    # ──────────────────────────────────────────────────────────────────────────
    # Full lifecycle tests
    # ──────────────────────────────────────────────────────────────────────────

    def test_full_lifecycle_draft_to_confirmed(self):
        """Complete happy path: draft → process → confirm, all state/qty correct."""
        p = self._make_product('TST-FL01')
        po = self._make_po(p, 10.0, price_unit=5.0)
        po_line = po.order_line[0]
        self.assertAlmostEqual(po_line.qty_open, 10.0)

        imp = self._make_import([('TST-FL01', 10, 5.0)])
        self.assertEqual(imp.state, 'draft')
        self.assertFalse(imp.line_ids)

        imp.action_process_file()
        self.assertEqual(imp.state, 'processed')
        self.assertEqual(len(imp.line_ids), 1)
        self.assertAlmostEqual(imp.line_ids[0].qty_split, 10.0)

        imp.action_confirm_delivery()
        self.assertEqual(imp.state, 'confirmed')
        po_line.invalidate_recordset()
        self.assertAlmostEqual(po_line.qty_split, 10.0)
        self.assertAlmostEqual(po_line.qty_open, 0.0)

    def test_full_lifecycle_confirm_reset_reprocess(self):
        """Confirm → reset-to-processed → reset-to-draft → reprocess."""
        p = self._make_product('TST-FL02')
        po = self._make_po(p, 10.0)
        po_line = po.order_line[0]

        imp = self._make_import([('TST-FL02', 6, 10.0)])
        imp.action_process_file()
        imp.action_confirm_delivery()
        po_line.invalidate_recordset()
        self.assertAlmostEqual(po_line.qty_split, 6.0)

        imp.action_reset_to_processed()
        po_line.invalidate_recordset()
        self.assertAlmostEqual(po_line.qty_split, 0.0)
        self.assertEqual(imp.state, 'processed')

        imp.action_reset_to_draft()
        self.assertEqual(imp.state, 'draft')
        self.assertFalse(imp.line_ids)

        imp.write({'file_data': self._make_excel([('TST-FL02', 4, 10.0)])})
        imp.action_process_file()
        imp.action_confirm_delivery()
        po_line.invalidate_recordset()
        self.assertAlmostEqual(po_line.qty_split, 4.0)

    def test_two_sequential_imports_deplete_qty_open(self):
        """Second import sees reduced open qty after first is confirmed."""
        p = self._make_product('TST-FL03')
        po = self._make_po(p, 10.0)
        po_line = po.order_line[0]

        # First import: 6 units
        imp1 = self._make_import([('TST-FL03', 6, 10.0)])
        imp1.action_process_file()
        imp1.action_confirm_delivery()
        po_line.invalidate_recordset()
        self.assertAlmostEqual(po_line.qty_open, 4.0)

        # Second import: only 4 left
        imp2 = self._make_import([('TST-FL03', 4, 10.0)])
        imp2.action_process_file()
        self.assertAlmostEqual(imp2.line_ids[0].qty_split, 4.0)
        imp2.action_confirm_delivery()
        po_line.invalidate_recordset()
        self.assertAlmostEqual(po_line.qty_open, 0.0)

    def test_sequence_reference_generated_on_create(self):
        """A sequence reference (not 'New') is automatically assigned at create."""
        imp = self._make_import()
        self.assertTrue(imp.name)
        self.assertNotEqual(imp.name, 'New')

    def test_two_products_same_import_independent_splits(self):
        """Two different SKUs in one Excel file are split independently."""
        p1 = self._make_product('TST-FL04A')
        p2 = self._make_product('TST-FL04B')
        po = self.env['purchase.order'].create({
            'partner_id': self.vendor.id,
            'order_line': [
                Command.create({'product_id': p1.id, 'product_qty': 10.0, 'price_unit': 5.0}),
                Command.create({'product_id': p2.id, 'product_qty': 10.0, 'price_unit': 8.0}),
            ],
        })
        po.button_confirm()
        imp = self._process([('TST-FL04A', 3, 5.0), ('TST-FL04B', 7, 8.0)])
        self.assertEqual(imp.total_po_lines, 2)
        self.assertEqual(imp.total_received, 10)  # 3 + 7

        line_a = imp.line_ids.filtered(lambda l: l.product_id == p1)
        line_b = imp.line_ids.filtered(lambda l: l.product_id == p2)
        self.assertAlmostEqual(line_a[0].qty_split, 3.0)
        self.assertAlmostEqual(line_b[0].qty_split, 7.0)

    # ──────────────────────────────────────────────────────────────────────────
    # Compute — qty_open double-counting prevention (Purchase Order Line)
    # ──────────────────────────────────────────────────────────────────────────

    def test_po_line_compute_qty_open_scenarios(self):
        """ Test various combinations of qty_received and qty_split to ensure qty_open calculates correctly without double-counting. """
        p = self._make_product('TST-QTY-OPEN')
        po = self._make_po(p, 10.0)
        po_line = po.order_line[0]

        # Scenario 1: Items allocated via split, none received yet
        po_line.write({'qty_split': 4.0, 'qty_received': 0.0})
        self.assertAlmostEqual(po_line.qty_open, 6.0, 
                         msg="Ordered 10, Split 4, Received 0 -> Open should be 6")

        # Scenario 2: The Bug Fix - Items allocated and those same items received
        po_line.write({'qty_split': 4.0, 'qty_received': 4.0})
        self.assertAlmostEqual(po_line.qty_open, 6.0, 
                         msg="Ordered 10, Split 4, Received 4 -> Open should still be 6 (no double counting)")

        # Scenario 3: Received more items than were explicitly split/allocated
        po_line.write({'qty_split': 4.0, 'qty_received': 6.0})
        self.assertAlmostEqual(po_line.qty_open, 4.0, 
                         msg="Ordered 10, max(Split 4, Received 6) is 6 -> Open should be 4")

        # Scenario 4: Over-allocated (prevents negative open quantity)
        po_line.write({'qty_split': 12.0, 'qty_received': 0.0})
        self.assertAlmostEqual(po_line.qty_open, 0.0, 
                         msg="Open quantity should not drop below 0.0 when split > ordered")

        # Scenario 5: Over-received (prevents negative open quantity)
        po_line.write({'qty_split': 0.0, 'qty_received': 15.0})
        self.assertAlmostEqual(po_line.qty_open, 0.0, 
                         msg="Open quantity should not drop below 0.0 when received > ordered")
