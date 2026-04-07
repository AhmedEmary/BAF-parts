import base64
import io
import openpyxl
from odoo.tests import TransactionCase, tagged
from odoo.exceptions import UserError
from odoo import Command

@tagged('post_install', '-at_install')
class TestGroupedPOExport(TransactionCase):

    def setUp(self):
        super().setUp()
        # 1. Setup Vendors (Trusted vs Untrusted)
        self.vendor_trusted = self.env['res.partner'].create({
            'name': 'Trusted Vendor',
            'is_trusted_vendor': True,
        })
        self.vendor_untrusted = self.env['res.partner'].create({
            'name': 'Untrusted Vendor',
            'is_trusted_vendor': False,
        })
        self.customer = self.env['res.partner'].create({'name': 'Test Customer'})
        
        # 2. Setup Product
        self.brand = self.env['product.brand'].create({'name': 'Test Brand'})
        self.product = self.env['product.product'].create({
            'name': 'Test Widget',
            'type': 'consu',
            'list_price': 100.0,
            'brand': self.brand.id,
            'default_code': 'SKU123'
        })

    def test_vendor_consistency_check(self):
        """ Test that selecting POs from different suppliers raises an error [cite: 1102] """
        po1 = self.env['purchase.order'].create({'partner_id': self.vendor_trusted.id})
        po2 = self.env['purchase.order'].create({'partner_id': self.vendor_untrusted.id})
        
        # Expect Error: "Different suppliers detected"
        with self.assertRaises(UserError):
            (po1 | po2).action_send_grouped_po_email()

    def test_export_trusted_vendor_columns(self):
        """ Test export for trusted vendor includes 'Customer' column [cite: 1107] """
        # Create SO linked to Customer
        so = self.env['sale.order'].create({
            'partner_id': self.customer.id,
            'order_line': [Command.create({'product_id': self.product.id})]
        })
        
        # Create PO linked to SO
        po = self.env['purchase.order'].create({
            'partner_id': self.vendor_trusted.id,
            'sale_order_id': so.id,
            'order_line': [Command.create({
                'product_id': self.product.id, 
                'product_qty': 1.0,
                'retail_price': 100.0, # Custom field test
                'price_unit': 80.0
            })]
        })

        # Run Action
        action = po.action_send_grouped_po_email()
        
        # [cite_start]1. Verify Status Update [cite: 1101]
        self.assertEqual(po.send_po_status, 'success', "PO status should be updated to 'success'")
        
        # 2. Verify Attachment Created
        attachment_id = action['context']['default_attachment_ids'][0]
        attachment = self.env['ir.attachment'].browse(attachment_id)
        self.assertTrue(attachment, "An Excel attachment should be generated")
        
        # 3. Verify Excel Content
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(attachment.datas)))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        # Trusted Vendor -> Customer column visible
        self.assertIn("Customer", headers, "Trusted Vendor export MUST include 'Customer' column")
        
        # Verify Data Mapping
        row_values = [cell.value for cell in ws[2]]
        self.assertIn(self.customer.name, row_values)
        self.assertIn('Test Brand', row_values)
        self.assertIn('SKU123', row_values)

    def test_export_untrusted_vendor_columns(self):
        """ Test export for untrusted vendor HIDES 'Customer' column [cite: 1107] """
        po = self.env['purchase.order'].create({
            'partner_id': self.vendor_untrusted.id,
            'order_line': [Command.create({'product_id': self.product.id, 'product_qty': 1.0})]
        })

        action = po.action_send_grouped_po_email()
        
        attachment_id = action['context']['default_attachment_ids'][0]
        attachment = self.env['ir.attachment'].browse(attachment_id)
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(attachment.datas)))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        # Untrusted Vendor -> No Customer column
        self.assertNotIn("Customer", headers, "Untrusted Vendor export MUST NOT include 'Customer' column")

    def test_dropship_sheet_separation(self):
        """ Test that standard and dropship orders go to different sheets [cite: 1103] """
        # Dropship Address
        dropship_addr = self.env['res.partner'].create({'name': 'Dropship Loc', 'street': '123 Drop St'})
        
        # Standard PO
        po_std = self.env['purchase.order'].create({
            'partner_id': self.vendor_trusted.id,
            'order_line': [Command.create({'product_id': self.product.id, 'product_qty': 1})]
        })
        
        # Dropship PO (has dest_address_id)
        po_drop = self.env['purchase.order'].create({
            'partner_id': self.vendor_trusted.id,
            'dest_address_id': dropship_addr.id,
            'order_line': [Command.create({'product_id': self.product.id, 'product_qty': 1})]
        })

        # Run action on both
        action = (po_std | po_drop).action_send_grouped_po_email()
        
        attachment_id = action['context']['default_attachment_ids'][0]
        attachment = self.env['ir.attachment'].browse(attachment_id)
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(attachment.datas)))
        
        # [cite_start]Verify Sheets [cite: 1103]
        self.assertIn("Standard Orders", wb.sheetnames)
        self.assertIn("Dropship Orders", wb.sheetnames)
        
        # [cite_start]Verify Dropship Headers have address info [cite: 1104]
        ws_drop = wb["Dropship Orders"]
        headers = [cell.value for cell in ws_drop[1]]
        self.assertIn("Delivery Address", headers)
