import base64
import io
import openpyxl

from odoo.tests import TransactionCase, tagged


def _xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue())


@tagged('post_install', '-at_install')
class TestDiscountLinePartnerScope(TransactionCase):

    def setUp(self):
        super().setUp()
        self.Disc = self.env['baf.discount.line']
        self.vendor_a = self.env['res.partner'].create({'name': 'Vendor A'})
        self.vendor_b = self.env['res.partner'].create({'name': 'Vendor B'})

    def test_purchase_row_scoped_to_partner(self):
        self.Disc.create({
            'table_type': 'purchase', 'column_key': 'ZZTEST_COL',
            'discount_code': 'ZZ10', 'discount_pct': 20.0,
            'partner_id': self.vendor_a.id,
        })
        # Vendor A sees its row
        self.assertEqual(
            self.Disc.get_discount_pct('purchase', 'ZZTEST_COL', 'ZZ10', partner=self.vendor_a),
            20.0,
        )
        # Vendor B has no row -> None (not 0.0)
        self.assertIsNone(
            self.Disc.get_discount_pct('purchase', 'ZZTEST_COL', 'ZZ10', partner=self.vendor_b),
        )

    def test_missing_row_returns_none(self):
        self.assertIsNone(
            self.Disc.get_discount_pct('purchase', 'ZZTEST_NOPE', 'ZZ10', partner=self.vendor_a),
        )

    def test_sales_row_unscoped(self):
        self.Disc.create({
            'table_type': 'sales', 'column_key': 'ZZTEST_SALES_GR1',
            'discount_code': 'ZZ11', 'discount_pct': 5.0,
        })
        self.assertEqual(
            self.Disc.get_discount_pct('sales', 'ZZTEST_SALES_GR1', 'ZZ11'),
            5.0,
        )


@tagged('post_install', '-at_install')
class TestVendorFields(TransactionCase):

    def test_vendor_method_and_sb(self):
        vendor = self.env['res.partner'].create({
            'name': 'Vendor M',
            'baf_purchase_method': 'matrix',
            'baf_sb_surcharge_pct': 5.2,
        })
        self.assertEqual(vendor.baf_purchase_method, 'matrix')
        self.assertEqual(vendor.baf_sb_surcharge_pct, 5.2)

    def test_supplier_code_removed(self):
        self.assertNotIn('baf_supplier_code', self.env['res.partner']._fields)

    def test_purchase_line_inverse(self):
        vendor = self.env['res.partner'].create({'name': 'Vendor I'})
        line = self.env['baf.discount.line'].create({
            'table_type': 'purchase', 'column_key': 'JLR',
            'discount_code': '1A', 'discount_pct': 30.0,
            'partner_id': vendor.id,
        })
        self.assertIn(line, vendor.baf_purchase_line_ids)


@tagged('post_install', '-at_install')
class TestProductFieldsRemoved(TransactionCase):

    def test_obsolete_fields_gone(self):
        fields_ = self.env['product.template']._fields
        for name in ('disc_code_1', 'disc_code_2', 'supplier_route',
                     'baf_sb_surcharge_override'):
            self.assertNotIn(name, fields_)

    def test_kept_fields_present(self):
        fields_ = self.env['product.template']._fields
        for name in ('baf_discount_code', 'baf_column_key', 'baf_mod'):
            self.assertIn(name, fields_)


@tagged('post_install', '-at_install')
class TestPurchaseEngine(TransactionCase):

    def setUp(self):
        super().setUp()
        Brand = self.env['product.brand']
        Partner = self.env['res.partner']
        Tmpl = self.env['product.template']
        self.Disc = self.env['baf.discount.line']

        self.brand_bmw = Brand.create({'name': 'BMW'})
        self.brand_jag = Brand.create({'name': 'Jaguar'})

        # UPE 100 everywhere for easy arithmetic
        self.prod_bmw = Tmpl.create({
            'name': 'BMW part', 'list_price': 100.0, 'brand': self.brand_bmw.id,
            'baf_discount_code': '10', 'baf_type_code': 1,  # -> BMW_T12
        })
        self.prod_bmw_sb = Tmpl.create({
            'name': 'BMW SB part', 'list_price': 100.0, 'brand': self.brand_bmw.id,
            'baf_discount_code': '10', 'baf_type_code': 1, 'baf_mod': 'sb',
        })
        self.prod_moto = Tmpl.create({
            'name': 'BMW moto part', 'list_price': 100.0, 'brand': self.brand_bmw.id,
            'baf_discount_code': '10', 'baf_type_code': 1, 'baf_mod': 'motorcycle',
        })
        self.prod_jag = Tmpl.create({
            'name': 'JLR part', 'list_price': 100.0, 'brand': self.brand_jag.id,
            'baf_discount_code': '1A',
        })

        self.vendor_matrix = Partner.create({
            'name': 'Matrix Vendor', 'baf_purchase_method': 'matrix',
            'baf_sb_surcharge_pct': 5.0,
        })
        self.vendor_codes = Partner.create({
            'name': 'Codes Vendor', 'baf_purchase_method': 'codes',
        })
        self.vendor_direct = Partner.create({
            'name': 'Direct Vendor', 'baf_purchase_method': 'direct',
        })

        self.Disc.create({
            'table_type': 'purchase', 'column_key': 'BMW_T12',
            'discount_code': '10', 'discount_pct': 20.0,
            'partner_id': self.vendor_matrix.id,
        })
        self.Disc.create({
            'table_type': 'purchase', 'column_key': 'MOTO',
            'discount_code': '10', 'discount_pct': 40.0,
            'partner_id': self.vendor_matrix.id,
        })
        code = self.env['discount.code'].create({'name': '10'})
        self.env['discount.code.value'].create({
            'code_id': code.id, 'partner_id': self.vendor_codes.id,
            'percentage': 30.0,
        })
        self.env['product.supplierinfo'].create({
            'partner_id': self.vendor_direct.id,
            'product_tmpl_id': self.prod_jag.id, 'price': 65.0,
        })

    def test_matrix_method(self):
        d = self.prod_bmw.baf_get_purchase_price_details(self.vendor_matrix)
        self.assertEqual(d['pricing_method'], 'matrix')
        self.assertEqual(d['price'], 80.0)
        self.assertEqual(d['column_key'], 'BMW_T12')

    def test_matrix_sb_extra_discount(self):
        d = self.prod_bmw_sb.baf_get_purchase_price_details(self.vendor_matrix)
        # 100 * 0.80 * 0.95 = 76.0
        self.assertEqual(d['price'], 76.0)
        self.assertEqual(d['sb_surcharge'], 5.0)

    def test_matrix_moto_column(self):
        d = self.prod_moto.baf_get_purchase_price_details(self.vendor_matrix)
        self.assertEqual(d['column_key'], 'MOTO')
        self.assertEqual(d['price'], 60.0)

    def test_matrix_miss_returns_none(self):
        # BMW product, but this vendor has no row for JLR code
        self.assertIsNone(
            self.prod_jag.baf_get_purchase_price_details(self.vendor_matrix))

    def test_codes_method(self):
        d = self.prod_bmw.baf_get_purchase_price_details(self.vendor_codes)
        self.assertEqual(d['pricing_method'], 'codes')
        self.assertEqual(d['price'], 70.0)

    def test_codes_miss_returns_none(self):
        self.assertIsNone(
            self.prod_jag.baf_get_purchase_price_details(self.vendor_codes))

    def test_direct_method(self):
        d = self.prod_jag.baf_get_purchase_price_details(self.vendor_direct)
        self.assertEqual(d['pricing_method'], 'direct')
        self.assertEqual(d['price'], 65.0)

    def test_direct_miss_returns_none(self):
        self.assertIsNone(
            self.prod_bmw.baf_get_purchase_price_details(self.vendor_direct))

    def test_no_method_returns_none(self):
        vendor = self.env['res.partner'].create({'name': 'No Method'})
        self.assertIsNone(
            self.prod_bmw.baf_get_purchase_price_details(vendor))


@tagged('post_install', '-at_install')
class TestPurchaseLinePrice(TransactionCase):

    def test_po_line_price_from_vendor_method(self):
        Brand = self.env['product.brand']
        Partner = self.env['res.partner']
        brand = Brand.create({'name': 'BMW'})
        vendor = Partner.create({
            'name': 'Matrix Vendor', 'baf_purchase_method': 'matrix',
            'baf_brand_ids': [(6, 0, [brand.id])],
        })
        prod = self.env['product.template'].create({
            'name': 'BMW part', 'list_price': 100.0, 'brand': brand.id,
            'baf_discount_code': '10', 'baf_type_code': 1,
        })
        self.env['baf.discount.line'].create({
            'table_type': 'purchase', 'column_key': 'BMW_T12',
            'discount_code': '10', 'discount_pct': 20.0,
            'partner_id': vendor.id,
        })
        po = self.env['purchase.order'].create({'partner_id': vendor.id})
        line = self.env['purchase.order.line'].new({
            'order_id': po.id, 'product_id': prod.product_variant_id.id,
        })
        line._onchange_product_id_custom()
        self.assertEqual(line.price_unit, 80.0)
        self.assertEqual(line.baf_column_key, 'BMW_T12')

    def test_po_line_price_survives_core_compute(self):
        # The stored core compute (_compute_price_unit_and_date_planned_and_name)
        # must not overwrite the BAF discounted price with standard cost.
        Brand = self.env['product.brand']
        Partner = self.env['res.partner']
        brand = Brand.create({'name': 'BMW'})
        vendor = Partner.create({
            'name': 'Matrix Vendor', 'baf_purchase_method': 'matrix',
            'baf_brand_ids': [(6, 0, [brand.id])],
        })
        prod = self.env['product.template'].create({
            'name': 'BMW part', 'list_price': 100.0, 'standard_price': 100.0,
            'brand': brand.id, 'baf_discount_code': '10', 'baf_type_code': 1,
        })
        self.env['baf.discount.line'].create({
            'table_type': 'purchase', 'column_key': 'BMW_T12',
            'discount_code': '10', 'discount_pct': 20.0,
            'partner_id': vendor.id,
        })
        po = self.env['purchase.order'].create({'partner_id': vendor.id})
        line = self.env['purchase.order.line'].create({
            'order_id': po.id, 'product_id': prod.product_variant_id.id,
            'product_qty': 5.0,
        })
        # Force the core compute to re-run (as a qty change would in the UI).
        line.product_qty = 7.0
        line.invalidate_recordset(['price_unit'])
        self.assertEqual(line.price_unit, 80.0)
        self.assertEqual(line.baf_discount_pct, 20.0)

    def test_po_line_reprices_when_vendor_changes(self):
        # Changing the PO vendor must reprice the line via the per-vendor engine.
        Brand = self.env['product.brand']
        Partner = self.env['res.partner']
        brand = Brand.create({'name': 'BMW'})
        vendor_a = Partner.create({
            'name': 'Vendor A', 'baf_purchase_method': 'matrix',
            'baf_brand_ids': [(6, 0, [brand.id])]})
        vendor_b = Partner.create({
            'name': 'Vendor B', 'baf_purchase_method': 'matrix',
            'baf_brand_ids': [(6, 0, [brand.id])]})
        prod = self.env['product.template'].create({
            'name': 'BMW part', 'list_price': 100.0, 'brand': brand.id,
            'baf_discount_code': '10', 'baf_type_code': 1})
        Disc = self.env['baf.discount.line']
        Disc.create({'table_type': 'purchase', 'column_key': 'BMW_T12',
                     'discount_code': '10', 'discount_pct': 20.0,
                     'partner_id': vendor_a.id})   # A -> 80
        Disc.create({'table_type': 'purchase', 'column_key': 'BMW_T12',
                     'discount_code': '10', 'discount_pct': 40.0,
                     'partner_id': vendor_b.id})   # B -> 60
        po = self.env['purchase.order'].create({'partner_id': vendor_a.id})
        line = self.env['purchase.order.line'].create({
            'order_id': po.id, 'product_id': prod.product_variant_id.id,
            'product_qty': 1.0})
        self.assertEqual(line.price_unit, 80.0)
        po.partner_id = vendor_b.id
        line.invalidate_recordset(['price_unit'])
        self.assertEqual(line.price_unit, 60.0)
        # Switch to a vendor with no discount for this product -> full retail.
        vendor_none = Partner.create({
            'name': 'Vendor None', 'baf_purchase_method': 'matrix',
            'baf_brand_ids': [(6, 0, [brand.id])]})  # no discount rows
        po.partner_id = vendor_none.id
        line.invalidate_recordset(['price_unit', 'baf_discount_pct'])
        self.assertEqual(line.price_unit, 100.0)   # back to retail (UPE)
        self.assertEqual(line.baf_discount_pct, 0.0)

    def test_po_line_retail_when_vendor_has_no_discount(self):
        # Selecting a no-discount vendor from the start -> unit price = retail.
        Brand = self.env['product.brand']
        Partner = self.env['res.partner']
        brand = Brand.create({'name': 'BMW'})
        vendor = Partner.create({
            'name': 'No-discount Vendor', 'baf_purchase_method': 'matrix',
            'baf_brand_ids': [(6, 0, [brand.id])]})  # no rows uploaded
        prod = self.env['product.template'].create({
            'name': 'BMW part', 'list_price': 100.0, 'brand': brand.id,
            'baf_discount_code': '10', 'baf_type_code': 1})
        po = self.env['purchase.order'].create({'partner_id': vendor.id})
        line = self.env['purchase.order.line'].create({
            'order_id': po.id, 'product_id': prod.product_variant_id.id,
            'product_qty': 1.0})
        self.assertEqual(line.price_unit, 100.0)
        self.assertEqual(line.baf_discount_pct, 0.0)


@tagged('post_install', '-at_install')
class TestVendorImport(TransactionCase):
    """Per-vendor pricing upload is now inline on res.partner
    (action_import_vendor_pricing_file), not a separate wizard."""

    def setUp(self):
        super().setUp()
        self.vendor = self.env['res.partner'].create({
            'name': 'Imp Vendor', 'baf_purchase_method': 'matrix',
        })
        self.Tmpl = self.env['product.template']
        self.Brand = self.env['product.brand']

    def _run(self, method, rows):
        self.vendor.baf_purchase_method = method
        self.vendor.baf_pricing_file = _xlsx(rows)
        self.vendor.baf_pricing_filename = 'f.xlsx'
        return self.vendor.action_import_vendor_pricing_file()

    def test_vendor_matrix_import(self):
        self._run('matrix', [
            ['DC', 'BMW_T12', 'MOTO'],
            ['10', '20', '40'],
            ['11', '15', ''],
        ])
        Disc = self.env['baf.discount.line']
        self.assertEqual(
            Disc.get_discount_pct('purchase', 'BMW_T12', '10', partner=self.vendor), 20.0)
        self.assertEqual(
            Disc.get_discount_pct('purchase', 'MOTO', '10', partner=self.vendor), 40.0)
        self.assertIsNone(
            Disc.get_discount_pct('purchase', 'MOTO', '11', partner=self.vendor))
        # Upload field cleared after processing.
        self.assertFalse(self.vendor.baf_pricing_file)

    def test_vendor_matrix_header_normalization(self):
        # Descriptive labels map to canonical baf_column_key (BMW/MINI/MOTO only).
        self._run('matrix', [
            ['DC', 'BMW TA 1-2-4-6-8', 'BMW TA 3-5-7-9', 'MINI TA  3-5-7-9'],
            ['10', '8', '9', '7'],
        ])
        Disc = self.env['baf.discount.line']
        self.assertEqual(
            Disc.get_discount_pct('purchase', 'BMW_T12', '10', partner=self.vendor), 8.0)
        self.assertEqual(
            Disc.get_discount_pct('purchase', 'BMW_T39', '10', partner=self.vendor), 9.0)
        self.assertEqual(
            Disc.get_discount_pct('purchase', 'MINI_T39', '10', partner=self.vendor), 7.0)
        self.assertIsNone(
            Disc.get_discount_pct('purchase', 'BMW TA 1-2-4-6-8', '10', partner=self.vendor))

    def test_vendor_codes_import(self):
        self._run('codes', [['CODE', 'PCT'], ['ZZ10', '30'], ['ZZ1A', '35']])
        val = self.env['discount.code.value'].search([
            ('partner_id', '=', self.vendor.id), ('code_id.name', '=', 'ZZ10')])
        self.assertEqual(val.percentage, 30.0)

    def test_vendor_direct_import_by_sku(self):
        brand = self.Brand.create({'name': 'BMW'})
        prod = self.Tmpl.create({
            'name': 'P', 'default_code': 'ZZDC1', 'sku': 'ZZSKU1',
            'brand': brand.id, 'list_price': 100.0})
        self._run('direct', [['SKU', 'Discounted Price'], ['ZZSKU1', '65']])
        si = self.env['product.supplierinfo'].search([
            ('partner_id', '=', self.vendor.id), ('product_tmpl_id', '=', prod.id)])
        self.assertEqual(si.price, 65.0)

    def test_vendor_direct_brand_disambiguation(self):
        bmw = self.Brand.create({'name': 'BMW'})
        mini = self.Brand.create({'name': 'MINI'})
        p_bmw = self.Tmpl.create({
            'name': 'A', 'default_code': 'ZZDCA', 'sku': 'ZZDUP1',
            'brand': bmw.id, 'list_price': 100.0})
        p_mini = self.Tmpl.create({
            'name': 'B', 'default_code': 'ZZDCB', 'sku': 'ZZDUP1',
            'brand': mini.id, 'list_price': 100.0})
        res = self._run('direct', [['SKU', 'Discounted Price', 'BRAND'], ['ZZDUP1', '65', 'MINI']])
        Seller = self.env['product.supplierinfo']
        self.assertEqual(
            Seller.search([('partner_id', '=', self.vendor.id),
                           ('product_tmpl_id', '=', p_mini.id)]).price, 65.0)
        self.assertFalse(
            Seller.search([('partner_id', '=', self.vendor.id),
                           ('product_tmpl_id', '=', p_bmw.id)]))
        self.assertEqual(res['params']['type'], 'success')

    def test_vendor_direct_ambiguous_sku_warns(self):
        bmw = self.Brand.create({'name': 'BMW'})
        mini = self.Brand.create({'name': 'MINI'})
        p_bmw = self.Tmpl.create({
            'name': 'A', 'default_code': 'ZZDCC', 'sku': 'ZZDUP2',
            'brand': bmw.id, 'list_price': 100.0})
        p_mini = self.Tmpl.create({
            'name': 'B', 'default_code': 'ZZDCD', 'sku': 'ZZDUP2',
            'brand': mini.id, 'list_price': 100.0})
        res = self._run('direct', [['SKU', 'Discounted Price'], ['ZZDUP2', '65']])
        Seller = self.env['product.supplierinfo']
        self.assertFalse(Seller.search([
            ('partner_id', '=', self.vendor.id),
            ('product_tmpl_id', 'in', [p_bmw.id, p_mini.id])]))
        self.assertEqual(res['params']['type'], 'warning')
        self.assertIn('ZZDUP2', res['params']['message'])

    def test_import_replaces_previous_data(self):
        # A second import fully replaces the vendor's data; old rows are gone.
        self._run('matrix', [['DC', 'BMW_T12'], ['10', '20']])
        self._run('matrix', [['DC', 'BMW_T12'], ['11', '30']])
        Disc = self.env['baf.discount.line']
        self.assertIsNone(
            Disc.get_discount_pct('purchase', 'BMW_T12', '10', partner=self.vendor))
        self.assertEqual(
            Disc.get_discount_pct('purchase', 'BMW_T12', '11', partner=self.vendor), 30.0)

    def test_switching_method_wipes_other_store(self):
        # Import matrix, then import codes -> matrix rows wiped, only codes kept.
        self._run('matrix', [['DC', 'BMW_T12'], ['10', '20']])
        self._run('codes', [['CODE', 'PCT'], ['ZZ10', '30']])
        self.assertFalse(self.vendor.baf_purchase_line_ids)
        self.assertTrue(self.vendor.baf_code_value_ids)

    def test_method_change_resets_staged_file(self):
        self.vendor.baf_pricing_file = _xlsx([['DC', 'BMW_T12'], ['10', '20']])
        self.vendor.baf_pricing_filename = 'f.xlsx'
        self.vendor.baf_purchase_method = 'codes'
        self.vendor._onchange_baf_purchase_method_reset_file()
        self.assertFalse(self.vendor.baf_pricing_file)

    def test_matrix_two_row_template_header(self):
        # Matrix template ships a 2-row header (types row + "RG/Discount in %").
        # The second header row must be ignored, real data still loads.
        self._run('matrix', [
            ['#', 'BMW TA 1-2-4-6-8', 'BMW TA 3-5-7-9'],
            ['RG', 'Discount in %', 'Discount in %'],
            ['10', '20', '25'],
        ])
        Disc = self.env['baf.discount.line']
        self.assertEqual(
            Disc.get_discount_pct('purchase', 'BMW_T12', '10', partner=self.vendor), 20.0)
        self.assertIsNone(
            Disc.get_discount_pct('purchase', 'BMW_T12', 'RG', partner=self.vendor))

    def test_download_template_action(self):
        for method, first_header in [
            ('direct', ['sku', 'discounted price', 'brand']),
            ('codes', ['dc', 'discount in %']),
            ('matrix', ['#', 'BMW TA 1-2-4-6-8']),
        ]:
            self.vendor.baf_purchase_method = method
            res = self.vendor.action_download_pricing_template()
            self.assertEqual(res['type'], 'ir.actions.act_url')
            att = self.env['ir.attachment'].search(
                [('res_id', '=', self.vendor.id),
                 ('name', '=', 'vendor_%s_template.xlsx' % method)], limit=1)
            self.assertTrue(att)
            wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(att.datas)))
            header = [c for c in next(wb.active.iter_rows(values_only=True))]
            self.assertEqual(header[:len(first_header)], first_header)
