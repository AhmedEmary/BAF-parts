from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
from io import BytesIO
import openpyxl
from openpyxl.styles import Font


class SaleOrder(models.Model):
    _inherit = 'sale.order'

    b2b_so = fields.Char(string='B2B SO')
    shipping_method = fields.Selection(
        selection=[
            ('air', 'Air'),
            ('ocean', 'Ocean'),
            ('road', 'Road'),
            ('postal', 'Postal'),
            ('other', 'Other'),
        ],
        string='Shipping Method',
    )
    customer_po = fields.Char(string='Customer PO Number')

    coverage_percentage = fields.Float(
        string='Coverage %',
        compute='_compute_coverage_percentage',
        store=True
    )

    purchase_ids = fields.One2many(
        'purchase.order',
        'sale_order_id',
        string='Purchase Orders'
    )

    purchase_count = fields.Integer(
        string='Purchase Orders Count',
        compute='_compute_purchase_count'
    )

    pallet_count = fields.Integer(string='Pallets', compute='_compute_pallet_count')
    amount_delivered = fields.Monetary(
        string="Shipped",
        compute='_compute_amount_delivered',
        store=True,
        currency_field='currency_id',
    )

    @api.depends('order_line.qty_delivered', 'order_line.price_unit', 'order_line.product_uom_qty')
    def _compute_amount_delivered(self):
        for order in self:
            order.amount_delivered = sum(
                line.qty_delivered * line.price_unit for line in order.order_line
            )

    def _compute_pallet_count(self):
        for order in self:
            if not isinstance(order.id, int):
                order.pallet_count = 0
                continue

            pallets = self.env['warehouse.pallet'].search([
                ('line_ids.sale_order_id', '=', order.id)
            ])
            order.pallet_count = len(pallets)

    def action_view_pallets(self):
        self.ensure_one()
        pallets = self.env['warehouse.pallet'].search([
            ('line_ids.sale_order_id', '=', self.id)
        ])
        return {
            'name': 'Related Pallets',
            'type': 'ir.actions.act_window',
            'res_model': 'warehouse.pallet',
            'view_mode': 'list,form',
            'domain': [('id', 'in', pallets.ids)],
        }

    @api.depends('purchase_ids')
    def _compute_purchase_count(self):
        for order in self:
            order.purchase_count = len(order.purchase_ids)

    def action_view_purchase_orders(self):
        """ Smart button action to open linked Purchase Orders """
        self.ensure_one()
        action = self.env["ir.actions.actions"]._for_xml_id("purchase.purchase_form_action")
        action['domain'] = [('id', 'in', self.purchase_ids.ids)]
        action['context'] = {'default_sale_order_id': self.id}
        return action

    @api.depends('order_line.reserved_qty', 'order_line.purchased_qty', 'order_line.product_uom_qty', 'order_line.reserve_qty')
    def _compute_coverage_percentage(self):
        for order in self:
            total_covered = sum((line.reserved_qty if line.reserve_qty else 0.0) + line.purchased_qty
                                for line in order.order_line)
            total_needed = sum(line.product_uom_qty for line in order.order_line)

            if total_needed > 0:
                order.coverage_percentage = min(100.0, (total_covered / total_needed) * 100)
            else:
                order.coverage_percentage = 0.0

    def action_create_purchase_order(self):
        """ Create Purchase Orders for all lines in this SO that need stock """
        self.ensure_one()

        lines_to_process = self.order_line.filtered(lambda l: l.qty_to_purchase > 0)

        if not lines_to_process:
            raise UserError("Selected lines have no shortage to purchase.")

        grouped_lines = {}
        no_vendor_lines = []

        for line in lines_to_process:
            seller = line.product_id._select_seller(quantity=line.qty_to_purchase, uom_id=line.product_uom_id)

            if not seller:
                no_vendor_lines.append(line.product_id.display_name)
                continue

            vendor = seller.partner_id
            if vendor not in grouped_lines:
                grouped_lines[vendor] = []
            grouped_lines[vendor].append((line, seller))

        if no_vendor_lines:
            raise UserError(f"No vendor found for: {', '.join(no_vendor_lines)}")

        created_pos = self.env['purchase.order']
        for vendor, items in grouped_lines.items():
            po = self.env['purchase.order'].create({
                'partner_id': vendor.id,
                'origin': self.name,
                'company_id': self.company_id.id,
                'date_order': fields.Datetime.now(),
                'sale_order_id': self.id,
            })
            created_pos += po

            for (so_line, seller) in items:
                self.env['purchase.order.line'].create({
                    'order_id': po.id,
                    'product_id': so_line.product_id.id,
                    'name': so_line.name,
                    'product_qty': so_line.qty_to_purchase,
                    'product_uom_id': so_line.product_uom_id.id,
                    'retail_price': so_line.price_unit,
                    'date_planned': fields.Datetime.now(),
                })

        return {
            'name': 'Created Purchase Orders',
            'type': 'ir.actions.act_window',
            'res_model': 'purchase.order',
            'view_mode': 'tree,form',
            'domain': [('id', 'in', created_pos.ids)],
        }

    def update_all_sale_line_reserved_qty(self):
        """
        Updates the reserved quantity for all lines in the order.
        Reserves the minimum of (Stock Available) and (Quantity Needed).
        """
        for order in self:
            for line in order.order_line:
                line.invalidate_recordset(['purchased_qty'])

                line._compute_purchased_qty()
                if line.product_id and line.product_uom_qty > 0:
                    line.reserve_qty = True
                    available = line.stock_quantity
                    needed = line.product_uom_qty
                    line.reserved_qty = max(0, min(needed, available))
        return True

    @api.model
    def get_import_templates(self):
        return [{
            'label': self.env._('Import Template for Intelliwise Quotations'),
            'template': '/general_system_custom/static/xls/quotations_import_template.xlsx',
        }]

    def action_export_so_excel(self):
        """Generates an Excel file for selected Sales Orders with lines and totals."""
        if not openpyxl:
            raise UserError("The 'openpyxl' library is missing.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Orders"
        bold_font = Font(bold=True)

        # Define the headers
        headers = [
            "Order Reference", "Date", "Customer", "B2B SO",
            "Customer PO", "Shipping Method", "SKU", "Product",
            "Quantity", "Unit Price", "Subtotal"
        ]
        ws.append(headers)

        # Make headers bold
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_num).font = bold_font

        # Iterate through selected records and their lines
        for order in self:
            shipping_val = dict(order._fields['shipping_method'].selection).get(order.shipping_method) if order.shipping_method else ""

            for line in order.order_line:
                ws.append([
                    order.name or "",
                    str(order.date_order or ""),
                    order.partner_id.name or "",
                    order.b2b_so or "",
                    order.customer_po or "",
                    shipping_val,
                    line.product_id.default_code or "",
                    line.product_id.name or "",
                    line.product_uom_qty,
                    line.price_unit,
                    line.price_subtotal
                ])

            # Append Totals for this Order
            ws.append(["", "", "", "", "", "", "", "", "", "Untaxed Amount:", order.amount_untaxed])
            ws.append(["", "", "", "", "", "", "", "", "", "Taxes:", order.amount_tax])
            ws.append(["", "", "", "", "", "", "", "", "", "Total:", order.amount_total])

            # Make the totals bold
            for row_idx in range(ws.max_row - 2, ws.max_row + 1):
                ws.cell(row=row_idx, column=10).font = bold_font
                ws.cell(row=row_idx, column=11).font = bold_font

            # Empty row as a separator if multiple orders are exported
            ws.append([])

        # Save to buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = base64.b64encode(output.read())
        output.close()

        attachment = self.env['ir.attachment'].create({
            'name': f'Sales_Orders_Export_{fields.Date.today()}.xlsx',
            'type': 'binary',
            'datas': file_content,
            'res_model': 'sale.order',
            'res_id': self.ids[0] if self.ids else False,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_export_transaction_report_excel(self):
        """ Generates a detailed transaction report (Sales & Refunds) for selected Sales Orders """
        if not openpyxl:
            raise UserError("The 'openpyxl' library is missing.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transaction Report"
        bold_font = Font(bold=True)

        headers = [
            "NR.Ordine SP", "Mercato", "Numero transaz.", "Tipo Transaz", "Vettore",
            "Codice di reso", "Data", "IVA", "SELLER_SKU", "Codice OEM", "Descrizione",
            "Totale", "P.IVA Cliente", "Destinazione merce", "TAXABLE_JURISDICTION",
            "EXPORT_OUTSIDE_EU", "NR.Fattura F.LE", "Nome cliente", "Respons. IVA",
            "Con fattura", "NOTA", "UE", "mese", "Data di creazione", "Creatore",
            "Ultimo modifica", "Autore della modifica"
        ]
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_num).font = bold_font

        for order in self:
            # If no invoices exist, we might still want to report the SO as a SALE without invoice.
            # But normally, this report relies on the posted financial documents (Invoices/Refunds)
            moves = order.invoice_ids.filtered(lambda m: m.state == 'posted')

            if not moves:
                moves = [None] # Dummy iteration if no invoice is found but we still want to report the order

            for move in moves:
                trans_type = "SALE"
                date_val = ""
                month_val = ""
                inv_number = ""

                if move:
                    if move.move_type == 'out_refund':
                        trans_type = "REFUND"
                    elif move.move_type in ('out_invoice', 'out_receipt'):
                        trans_type = "SALE"

                    date_val = move.invoice_date.strftime('%d/%m/%y') if move.invoice_date else ""
                    month_val = move.invoice_date.strftime('%m') if move.invoice_date else ""
                    inv_number = move.name
                else:
                    date_val = order.date_order.strftime('%d/%m/%y') if order.date_order else ""
                    month_val = order.date_order.strftime('%m') if order.date_order else ""

                # Determine Market from tracking refs or tags (Customize as needed)
                mercato = "Shopify"
                if order.customer_po and 'ebay' in order.customer_po.lower():
                    mercato = "eBay"
                elif order.customer_po and 'amazon' in order.customer_po.lower():
                    mercato = "Amazon"

                # Destination Country and EU checks
                dest_country = order.partner_shipping_id.country_id
                dest_code = dest_country.code if dest_country else ""
                eu_countries = self.env.ref('base.europe').country_ids.mapped('code')
                is_eu = "SI" if dest_code in eu_countries else "NO"
                export_outside_eu = "SI" if dest_code and dest_code not in eu_countries else "NO"

                for line in order.order_line.filtered(lambda l: not l.display_type):
                    # Multiplier for refund amounts
                    multiplier = -1 if trans_type == "REFUND" else 1

                    row_data = [
                        order.name or "",                                       # NR.Ordine SP
                        mercato,                                                # Mercato
                        order.customer_po or "",                                # Numero transaz.
                        trans_type,                                             # Tipo Transaz
                        dict(order._fields['shipping_method'].selection).get(order.shipping_method) or "", # Vettore
                        "",                                                     # Codice di reso (customize if you have a return model)
                        date_val,                                               # Data
                        "",                                                     # IVA (add tax % if needed: line.tax_id...)
                        line.product_id.sku or "",                              # SELLER_SKU
                        line.product_id.barcode or "",                          # Codice OEM (assuming barcode or specific OEM field)
                        line.product_id.name or line.name or "",                # Descrizione
                        (line.price_total * multiplier),                        # Totale
                        order.partner_id.vat or "",                             # P.IVA Cliente
                        dest_code,                                              # Destinazione merce
                        dest_code,                                              # TAXABLE_JURISDICTION
                        export_outside_eu,                                      # EXPORT_OUTSIDE_EU
                        inv_number,                                             # NR.Fattura F.LE
                        order.partner_id.name or "",                            # Nome cliente
                        "IT",                                                   # Respons. IVA (Default IT based on example)
                        "SI" if move else "NO",                                 # Con fattura
                        order.note or "",                                       # NOTA
                        is_eu,                                                  # UE
                        month_val,                                              # mese
                        order.create_date.strftime('%d/%m/%y') if order.create_date else "", # Data di creazione
                        order.create_uid.name or "",                            # Creatore
                        order.write_date.strftime('%d/%m/%y') if order.write_date else "",   # Ultimo modifica
                        order.write_uid.name or ""                              # Autore della modifica
                    ]
                    ws.append(row_data)

        # Output to buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = base64.b64encode(output.read())
        output.close()

        attachment = self.env['ir.attachment'].create({
            'name': f'Transaction_Report_{fields.Date.today()}.xlsx',
            'type': 'binary',
            'datas': file_content,
            'res_model': 'sale.order',
            'res_id': self.ids[0] if self.ids else False,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def _prepare_invoice(self):
        invoice_vals = super()._prepare_invoice()
        partner = self.partner_id
        eu_country_codes = set(self.env.ref('base.europe').country_ids.mapped('code'))

        if partner.country_code in eu_country_codes and not partner.vat and not partner.l10n_it_codice_fiscale:
            invoice_vals['move_type'] = 'out_receipt'
            receipt_group = self.env.ref('account.group_sale_receipts', raise_if_not_found=False)
            if receipt_group:
                base_group = self.env.ref('base.group_user')
                if receipt_group not in base_group.implied_ids:
                    base_group.sudo().write({'implied_ids': [(4, receipt_group.id)]})

                if self.env.user.id not in receipt_group.all_user_ids.ids:
                    receipt_group.sudo().write({'users': [(4, self.env.user.id)]})

        return invoice_vals
