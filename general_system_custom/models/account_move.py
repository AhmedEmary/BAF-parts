import base64
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    openpyxl = None
    Font = None


class AccountMoveLine(models.Model):
    _inherit = 'account.move.line'

    origin_country_id = fields.Many2one('res.country', related='product_id.origin', string="Origin Country", store=True, readonly=True)
    hs_code = fields.Char(related='product_id.hs_code', string="HS Code", store=True, readonly=True)
    supplier_delivery_ref = fields.Char(string="Delivery N.")
    linked_so_id = fields.Many2one('sale.order', string="Linked SO", readonly=True)
    linked_po_id = fields.Many2one('purchase.order', string="Linked PO", readonly=True)

    is_down_payment = fields.Boolean(string="Is Down Payment", default=False, copy=False)
    down_payment_origin_id = fields.Many2one(
        'account.move',
        string="Origin Down Payment Invoice",
        copy=False,
        index=True
    )


class AccountMove(models.Model):
    _inherit = 'account.move'

    pallet_ids = fields.One2many('warehouse.pallet', 'invoice_id', string="Source Pallets")
    pallet_count = fields.Integer(string='Pallet Count', compute='_compute_pallet_count')

    down_payment_usage_ids = fields.One2many(
        'account.move.line',
        'down_payment_origin_id',
        string="Down Payment Deductions"
    )

    amount_available_to_draw = fields.Monetary(
        string="Available DP Balance",
        compute='_compute_available_to_draw',
        currency_field='currency_id'
    )

    is_down_payment_invoice = fields.Boolean(
        string="Is Down Payment Invoice",
        default=False,
        copy=False,
        help="Check this box if this invoice represents a Down Payment."
    )

    @api.model_create_multi
    def create(self, vals_list):
        try:
            eu_country_codes = set(self.env.ref('base.europe').country_ids.mapped('code'))
        except ValueError:
            eu_country_codes = set()

        for vals in vals_list:
            if vals.get('move_type') == 'out_invoice' and vals.get('partner_id'):
                partner = self.env['res.partner'].browse(vals['partner_id'])
                if partner.country_id.code in eu_country_codes and not partner.vat and not getattr(partner, 'l10n_it_codice_fiscale', False):
                    vals['move_type'] = 'out_receipt'

                    # Ensure the current user has access to Sale Receipts
                    receipt_group = self.env.ref('account.group_sale_receipts', raise_if_not_found=False)
                    if receipt_group:
                        base_group = self.env.ref('base.group_user', raise_if_not_found=False)
                        if base_group and receipt_group not in base_group.implied_ids:
                            base_group.sudo().write({'implied_ids': [(4, receipt_group.id)]})

                        if self.env.user.id not in receipt_group.all_user_ids.ids:
                            receipt_group.sudo().write({'users': [(4, self.env.user.id)]})

        moves = super(AccountMove, self).create(vals_list)

        return moves

    @api.depends('pallet_ids')
    def _compute_pallet_count(self):
        for move in self:
            move.pallet_count = len(move.pallet_ids)

    @api.depends('amount_total', 'state', 'down_payment_usage_ids.price_subtotal', 'down_payment_usage_ids.parent_state')
    def _compute_available_to_draw(self):
        for move in self:
            if move.move_type == 'out_invoice' and move.state == 'posted':

                valid_usages = move.down_payment_usage_ids.filtered(
                    lambda l: l.parent_state != 'cancel'
                )

                total_used = sum(valid_usages.mapped('price_subtotal'))
                move.amount_available_to_draw = move.amount_total + total_used
            else:
                move.amount_available_to_draw = move.amount_total

    def action_view_pallets(self):
        self.ensure_one()
        return {
            'name': 'Source Pallets',
            'type': 'ir.actions.act_window',
            'res_model': 'warehouse.pallet',
            'view_mode': 'list,form',
            'domain': [('id', 'in', self.pallet_ids.ids)],
        }

    def action_post(self):
        """
         Confirmation & Validation
        - Generates official Number (Standard Odoo)
        - The link to Pallulet (Shipment List) is already established via 'pallet_ids'
        """
        res = super(AccountMove, self).action_post()
        return res

    def action_open_draw_down_payment(self):
        self.ensure_one()
        return {
            'name': 'Draw Down Payments',
            'type': 'ir.actions.act_window',
            'res_model': 'account.draw.down.payment.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_invoice_id': self.id},
        }

    def action_export_invoice_excel(self):
        """Generates an Excel file for selected Invoices/Bills with lines and totals."""
        if not openpyxl:
            raise UserError("The 'openpyxl' library is missing.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"
        bold_font = Font(bold=True)

        headers = [
            "Invoice Number", "Date", "Partner", "Type", "Status",
            "SKU", "Product", "Quantity", "Unit Price", "Subtotal",
            "Origin Country", "HS Code", "Supplier Delivery Ref"
        ]
        ws.append(headers)

        # Make headers bold
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_num).font = bold_font

        for move in self:
            move_type_val = dict(move._fields['move_type'].selection).get(move.move_type) or ""
            state_val = dict(move._fields['state'].selection).get(move.state) or ""

            # Only export actual product/service lines, ignoring layout/notes
            for line in move.invoice_line_ids.filtered(lambda l: l.display_type == 'product'):
                ws.append([
                    move.name or "",
                    str(move.invoice_date or ""),
                    move.partner_id.name or "",
                    move_type_val,
                    state_val,
                    line.product_id.default_code or "",
                    line.product_id.name or "",
                    line.quantity,
                    line.price_unit,
                    line.price_subtotal,
                    line.origin_country_id.name if line.origin_country_id else "",
                    line.hs_code or "",
                    line.supplier_delivery_ref or ""
                ])

            # Append Totals for this Invoice (Aligning 'Untaxed Amount' under 'Unit Price', and values under 'Subtotal')
            ws.append(["", "", "", "", "", "", "", "", "Untaxed Amount:", move.amount_untaxed, "", "", ""])
            ws.append(["", "", "", "", "", "", "", "", "Taxes:", move.amount_tax, "", "", ""])
            ws.append(["", "", "", "", "", "", "", "", "Total:", move.amount_total, "", "", ""])

            # Make the totals bold
            for row_idx in range(ws.max_row - 2, ws.max_row + 1):
                ws.cell(row=row_idx, column=9).font = bold_font
                ws.cell(row=row_idx, column=10).font = bold_font

            # Empty row as a separator if multiple invoices are exported
            ws.append([])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = base64.b64encode(output.read())
        output.close()

        attachment = self.env['ir.attachment'].create({
            'name': f'Invoices_Export_{fields.Date.today()}.xlsx',
            'type': 'binary',
            'datas': file_content,
            'res_model': 'account.move',
            'res_id': self.ids[0] if self.ids else False,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def _get_starting_sequence(self):
        # 1. Ensure the journal has the setting enabled before applying the REC/ prefix
        if self.company_id.account_fiscal_country_id.code == "IT" and self.move_type == 'out_receipt' and self.journal_id.type == 'sale':
            return "REC/%04d/0000" % self.date.year
        return super()._get_starting_sequence()

    def _get_last_sequence_domain(self, relaxed=False):
        """
        2. Force Odoo to look only at previous 'out_receipt' documents when calculating
        the next number for a receipt, and only 'out_invoice' when calculating an invoice.
        """
        where_string, param = super()._get_last_sequence_domain(relaxed)

        if self.journal_id.type == 'sale' and self.company_id.country_id.code == "IT":
            if self.move_type == 'out_receipt':
                where_string += " AND move_type = 'out_receipt'"
            elif self.move_type == 'out_invoice':
                where_string += " AND move_type = 'out_invoice'"

        return where_string, param

    def _post(self, soft=True):
        """
        Override _post to auto-generate and attach the PDF for Corrispettivi (out_receipt)
        upon validation.
        """
        posted = super()._post(soft=soft)

        it_receipts = posted.filtered(lambda m: m.move_type == 'out_receipt' and m.company_id.account_fiscal_country_id.code == 'IT' and not m.message_main_attachment_id)

        for move in it_receipts:
            pdf_content, _ = self.env['ir.actions.report']._render_qweb_pdf(
                'account.account_invoices',
                res_ids=move.ids
            )

            safe_name = move.name.replace('/', '_') if move.name else 'Receipt'
            attachment_name = f"{safe_name}.pdf"

            attachment = self.env['ir.attachment'].create({
                'name': attachment_name,
                'type': 'binary',
                'raw': pdf_content,
                'res_model': 'account.move',
                'res_id': move.id,
                'mimetype': 'application/pdf',
            })

            move.with_context(no_new_invoice=True).message_main_attachment_id = attachment.id

        return posted

