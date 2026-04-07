import logging
import base64
import re
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None

_logger = logging.getLogger(__name__)

class PurchaseOrder(models.Model):
    _inherit = 'purchase.order'

    sale_order_id = fields.Many2one(
        'sale.order', 
        string="Sales Order", 
        readonly=True,
        help="The Sales Order that triggered this Purchase Order."
    )

    line_count = fields.Integer(string='Line Count', compute='_compute_line_count')

    send_po_status = fields.Selection(
        selection=[('pending', 'Pending'), ('success', 'Successed')],
        string='Send PO',
        default='pending',
        copy=False,
        help="Status of the PO email sending process."
    )
    
    pallet_count = fields.Integer(string='Pallets', compute='_compute_pallet_count')

    def _compute_pallet_count(self):
        for order in self:
            if not isinstance(order.id, int):
                order.pallet_count = 0
                continue
                
            pallets = self.env['warehouse.pallet'].search([
                ('line_ids.purchase_order_id', '=', order.id)
            ])
            order.pallet_count = len(pallets)

    def action_view_pallets(self):
        self.ensure_one()
        pallets = self.env['warehouse.pallet'].search([
            ('line_ids.purchase_order_id', '=', self.id)
        ])
        return {
            'name': 'Related Pallets',
            'type': 'ir.actions.act_window',
            'res_model': 'warehouse.pallet',
            'view_mode': 'list,form',
            'domain': [('id', 'in', pallets.ids)],
        }

    @api.depends('order_line')
    def _compute_line_count(self):
        for order in self:
            order.line_count = len(order.order_line)

    def action_view_sale_order(self):
        self.ensure_one()
        return {
            'name': 'Sales Order',
            'type': 'ir.actions.act_window',
            'res_model': 'sale.order',
            'view_mode': 'form',
            'res_id': self.sale_order_id.id,
        }

    def action_view_po_lines(self):
        self.ensure_one()
        return {
            'name': f'Lines of {self.name}',
            'type': 'ir.actions.act_window',
            'res_model': 'purchase.order.line',
            'view_mode': 'list,form',
            'views': [
                (self.env.ref('intelliwise_custom.view_purchase_order_line_tree_intelliwise').id, 'list'),
                (False, 'form')
            ],
            'domain': [('order_id', '=', self.id)],
            'context': {'default_order_id': self.id},
        }

    def _sanitize(self, value):
        """ Helper to ensure False/None become empty strings for Excel """
        if value is False or value is None:
            return ""
        return value

    def action_send_grouped_po_email(self):
        """ 
        Reliable Excel generation with sanitization and error handling.
        """
        if not self:
            return
        
        partners = self.mapped('partner_id')
        if len(set(partners)) > 1:
            raise UserError(_("Different suppliers detected. Please select orders from a single supplier."))
        
        vendor = partners[0]
        is_trusted = vendor.is_trusted_vendor
        
        if not openpyxl:
            raise UserError(_("The 'openpyxl' library is missing."))

        wb = openpyxl.Workbook()
        
        headers = ["PO N."]
        if is_trusted:
            headers.append("Customer")
        
        headers.extend([
            "Brand", "SKU", "Quantity", "Retail", 
            "Discount 1", "Discount 2", "Surcharge", "Unit Net", "Total"
        ])
        
        ws_std = wb.active
        ws_std.title = "Standard Orders"
        ws_std.append(headers)
        
        ws_drop = wb.create_sheet("Dropship Orders")
        # Dropship headers have extra columns
        drop_headers = list(headers) + ["Delivery Address", "Phone"]
        ws_drop.append(drop_headers)

        has_std = False
        has_drop = False

        for po in self:
            is_dropship = bool(po.dest_address_id)
            
            customer_name = po.sale_order_id.partner_id.name if po.sale_order_id else ""
            
            for line in po.order_line:
                brand_name = self._sanitize(line.product_id.brand.name)
                sku = self._sanitize(line.product_id.default_code)
                
                row_data = [self._sanitize(po.name)]
                
                if is_trusted:
                    row_data.append(self._sanitize(customer_name))
                
                row_data.extend([
                    brand_name,
                    sku,
                    line.product_qty or 0.0,
                    line.retail_price or 0.0,
                    line.disc_code_1 or 0.0,
                    line.disc_code_2 or 0.0,
                    line.surcharge or 0.0,
                    line.price_unit or 0.0,
                    line.price_subtotal or 0.0
                ])

                if is_dropship:
                    addr = po.dest_address_id
                    if addr:
                        parts = [
                            addr.name,
                            addr.street,
                            addr.city,
                            addr.country_id.name
                        ]
                        address_str = ", ".join([str(p) for p in parts if p])
                        phone_val = addr.phone or ""
                    else:
                        address_str = ""
                        phone_val = ""

                    row_data.extend([address_str, phone_val])
                    ws_drop.append(row_data)
                    has_drop = True
                else:
                    ws_std.append(row_data)
                    has_std = True

        if not has_std and not has_drop:
            pass 
        else:
            if not has_drop and "Dropship Orders" in wb.sheetnames:
                del wb["Dropship Orders"]
            if not has_std and "Standard Orders" in wb.sheetnames:
                if len(wb.sheetnames) > 1:
                    del wb["Standard Orders"]
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = base64.b64encode(output.read())
        output.close()

        safe_vendor_name = re.sub(r'[\\/*?:"<>|]', "", vendor.name or "Vendor")
        attachment_name = f"Orders_{safe_vendor_name}_{fields.Date.today()}.xlsx"

        # 7. Create Attachment
        attachment = self.env['ir.attachment'].create({
            'name': attachment_name,
            'type': 'binary',
            'datas': file_content,
            'res_model': 'purchase.order',
            'res_id': self[0].id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        self.write({'send_po_status': 'success'})

        # 8. Open Composer
        template_id = self.env.ref('purchase.email_template_edi_purchase').id
        ctx = {
            'default_model': 'purchase.order',
            'default_res_ids': self.ids,
            'default_use_template': bool(template_id),
            'default_template_id': template_id,
            'default_attachment_ids': [attachment.id],
            'default_composition_mode': 'comment',
            'force_email': True,
        }
        return {
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'mail.compose.message',
            'views': [(False, 'form')],
            'view_id': False,
            'target': 'new',
            'context': ctx,
        }
