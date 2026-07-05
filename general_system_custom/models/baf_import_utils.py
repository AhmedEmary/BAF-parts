"""Shared spreadsheet-parsing helpers for BAF pricing imports.

Used by both the per-vendor import on res.partner and the (kept) global
sales-matrix wizard. No Odoo models here, so it is safe to import anywhere.
"""
import base64
import io
import re


def clean_cell(cell):
    if cell is None:
        return ""
    if isinstance(cell, float) and cell.is_integer():
        return str(int(cell))
    return str(cell).strip()


def parse_float(raw):
    s = str(raw).strip()
    if not s:
        return None
    try:
        return float(s.replace(',', '.'))
    except ValueError:
        return None


def read_workbook(file_name, file_data):
    """Parse an uploaded (base64) file into {sheet_name: [[cell, ...], ...]}.
    For CSV (single sheet) the key is ''. Raises on unreadable content."""
    file_name = (file_name or '').lower()
    file_content = base64.b64decode(file_data)
    sheets = {}
    if file_name.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
        for sn in wb.sheetnames:
            s = wb[sn]
            sheets[sn] = [[clean_cell(c) for c in row]
                          for row in s.iter_rows(values_only=True)]
    elif file_name.endswith('.xls'):
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_content)
        for sn in wb.sheet_names():
            s = wb.sheet_by_name(sn)
            sheets[sn] = [[clean_cell(c) for c in s.row_values(r)]
                          for r in range(s.nrows)]
    else:
        import csv
        csv_data = file_content.decode('utf-8-sig')
        sheets[''] = list(csv.reader(io.StringIO(csv_data), delimiter=','))
    return sheets


def first_sheet(sheets):
    """Return the first sheet's rows (per-vendor uploads are single-sheet)."""
    return next(iter(sheets.values()), [])


# BMW/MINI type-code buckets: 1,2,4,6,8 -> T12 ; 3,5,7,9 -> T39.
_T12_DIGITS = frozenset('12468')
_T39_DIGITS = frozenset('3579')


def normalize_matrix_header(raw):
    """Map a matrix column header to the canonical baf_column_key.

    Matrix pricing is BMW / MINI / MOTO only. Canonical keys pass through;
    descriptive labels are mapped, e.g.:
        'BMW TA 1-2-4-6-8' -> BMW_T12
        'MINI TA 3-5-7-9'  -> MINI_T39
        anything mentioning MOTO -> MOTO
    Unknown headers are returned uppercased with spaces as underscores.
    """
    s = re.sub(r'\s+', ' ', str(raw or '').strip().upper())
    if not s:
        return ''
    compact = s.replace(' ', '_')
    if re.fullmatch(r'(BMW|MINI)_T(12|39)', compact):
        return compact
    if 'MOTO' in compact:
        return 'MOTO'
    brand = 'BMW' if 'BMW' in s else ('MINI' if 'MINI' in s else None)
    if brand:
        digits = set(re.findall(r'\d', s))
        if 'T39' in compact or (digits and digits <= _T39_DIGITS):
            return f'{brand}_T39'
        return f'{brand}_T12'
    return compact
