import base64
import io
import re

import openpyxl
import xlrd
from odoo import models, fields, api, _
from odoo.exceptions import UserError

from .baf_product_pricing import (
    BAF_TYPE_BUCKETS,
    _normalize_brand,
    baf_brand_base_key,
)


class BafDiscountImportWizard(models.TransientModel):
    """Import a SALES discount matrix and build the matching sales groups.

    Only sales pricing is handled here; purchase pricing is per-vendor and
    uploaded inline on each contact (res.partner.action_import_vendor_pricing_file).

    You pick an import METHOD (the column layout), a BRAND FAMILY (whose brands
    set the column keys and the created group's scope), and a WRITE MODE (update
    the family's existing groups / add new ones, or delete them first and import
    only the sheet). Group columns are read from the sheet header, so any number
    of tiers works — a group is any column whose header names a tier:
        "SALE PRICE GR1" / "GR8" / "SALES GR3" -> GR1 / GR8 / GR3
        "GR_MOTORCYCLE" / anything with MOTO    -> MOTO
    """
    _name = 'baf.discount.import.wizard'
    _description = 'Import Sales Discount Matrix Tables'

    file_data = fields.Binary('File')
    file_name = fields.Char('File Name')

    import_method = fields.Selection([
        ('types_groups', 'Types & Groups Table (type sub-columns per group)'),
        ('groups_only', 'Groups Table (one column per group)'),
    ], string="Import Method", required=True, default='types_groups',
        help="Types & Groups Table: each group spans one column per (brand, "
             "type) — BMW_T12, BMW_T39, MINI_T12, MINI_T39 — for brands that "
             "price differently per type code (BMW/MINI).\n\n"
             "Groups Table: each group is a single column, shared by every brand "
             "of the family (JLR, Mercedes, ...).")

    family_id = fields.Many2one(
        'baf.brand.family',
        string="Brand Family",
        required=True,
        help="The brand family this sheet prices. Its brands set the discount "
             "column keys and the created group's scope, so they match what "
             "products of those brands look up. Merge brands into one family "
             "(e.g. Jaguar + Land Rover) to price them from one sheet.",
    )

    write_mode = fields.Selection([
        ('update', 'Update existing / add new groups only'),
        ('replace', 'Delete existing groups'),
    ], string="On Import", required=True, default='update',
        help="Update existing / add new groups only: keep every group already "
             "priced for this family; overwrite the codes present in the sheet "
             "and add new ones. Upload a single group (e.g. MOTO) without "
             "touching the others.\n\n"
             "Delete existing groups: first delete every existing sales discount "
             "line of this family, then import only what the sheet contains. "
             "Groups absent from the sheet are left with no pricing (their group "
             "records and customer assignments are kept).")

    @api.onchange('import_method')
    def _onchange_import_method_family(self):
        """Clear the family pick when switching layout, so the user re-chooses a
        family that fits the selected method."""
        self.family_id = False

    def _family_brands(self):
        return self.family_id.brand_ids

    def _brand_bases(self):
        """Distinct discount column bases of the family's brands, in brand order.
        Each base is the brand's own normalized name (brands are never merged
        into a shared base): Jaguar + Land Rover -> ['JAGUAR', 'LAND_ROVER'];
        BMW + MINI -> ['BMW', 'MINI']."""
        bases = []
        for brand in self._family_brands():
            base = baf_brand_base_key(brand.name)
            if base and base not in bases:
                bases.append(base)
        return bases

    def _is_type_split(self):
        """The chosen method decides the layout: Types & Groups splits each
        brand base into T12 / T39; Groups Table gives one column per base."""
        return self.import_method == 'types_groups'

    # ────────────────────────────────────────────────────────────────
    # Action entry points
    # ────────────────────────────────────────────────────────────────

    def action_import(self):
        if not self.file_data:
            raise UserError(_("Please upload a file."))

        sheets = self._read_workbook(self.file_name or '', self.file_data)
        rows = self._pick_sheet(sheets)
        counters = self._import_matrix(rows)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Discount Matrix Imported"),
                'message': _(
                    "%(c)d lines added, %(u)d updated, %(r)d cleared, "
                    "%(g)d groups created."
                ) % {'c': counters['created'], 'u': counters['updated'],
                     'r': counters['removed'], 'g': counters['groups_created']},
                'type': 'success',
                'next': {'type': 'ir.actions.act_window_close'},
            },
        }

    def action_download_template(self):
        """Download the Excel template for the selected family."""
        self.ensure_one()
        self._check_family()
        return {
            'type': 'ir.actions.act_url',
            'url': '/general_system_custom/discount_matrix_template'
                   '?family_id=%s&method=%s'
                   % (self.family_id.id, self.import_method or 'types_groups'),
            'target': 'self',
        }

    # ────────────────────────────────────────────────────────────────
    # File reading helpers
    # ────────────────────────────────────────────────────────────────

    @staticmethod
    def _clean(cell):
        if cell is None:
            return ""
        if isinstance(cell, float) and cell.is_integer():
            return str(int(cell))
        return str(cell).strip()

    def _read_workbook(self, file_name, file_data):
        """
        Parse the uploaded file into a dict {sheet_name: [[cell, ...], ...]}.
        For CSV (single sheet), returns {'': rows}.
        """
        file_name = file_name.lower()
        file_content = base64.b64decode(file_data)
        sheets = {}

        try:
            if file_name.endswith('.xlsx'):
                wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
                for sn in wb.sheetnames:
                    s = wb[sn]
                    sheets[sn] = [
                        [self._clean(c) for c in row]
                        for row in s.iter_rows(values_only=True)
                    ]
            elif file_name.endswith('.xls'):
                wb = xlrd.open_workbook(file_contents=file_content)
                for sn in wb.sheet_names():
                    s = wb.sheet_by_name(sn)
                    sheets[sn] = [
                        [self._clean(c) for c in s.row_values(r)]
                        for r in range(s.nrows)
                    ]
            else:
                import csv
                csv_data = file_content.decode('utf-8-sig')
                sheets[''] = list(csv.reader(io.StringIO(csv_data), delimiter=','))
        except Exception as e:
            raise UserError(
                _("Error reading file. Please ensure it is a valid CSV or Excel format. Details: %s")
                % str(e)
            )
        return sheets

    def _pick_sheet(self, sheets):
        """
        Pick the sheet holding the matrix: the first one naming any of the
        brand's column keys ('BMW-MINI-MOTORRAD' -> BMW, 'JLR' -> JLR),
        otherwise the first sheet in the workbook.
        """
        if not sheets:
            raise UserError(_("The uploaded file contains no sheets."))
        if len(sheets) > 1:
            bases = [b.replace('_', ' ') for b in self._brand_bases()]
            for sheet_name, rows in sheets.items():
                norm = _normalize_brand(sheet_name)
                if any(re.search(r'\b%s\b' % re.escape(b), norm) for b in bases):
                    return rows
        return next(iter(sheets.values()))

    # ────────────────────────────────────────────────────────────────
    # Helpers shared by importers
    # ────────────────────────────────────────────────────────────────

    @staticmethod
    def _parse_float(raw):
        s = str(raw).strip()
        if not s:
            return None
        try:
            return float(s.replace(',', '.'))
        except ValueError:
            return None

    @staticmethod
    def _group_suffix(cell):
        """Map a header cell to a group suffix, or None if it isn't a group.
        'SALE PRICE GR1'->GR1, 'GR8'->GR8, 'GR_MOTORCYCLE'/'MOTO'->MOTO."""
        text = str(cell or '').strip().upper()
        if not text:
            return None
        if 'MOTO' in text or 'MOTORRAD' in text or 'MOTORCYCLE' in text:
            return 'MOTO'
        m = re.search(r'GR[\s_]*0*(\d+)', text)
        return ('GR' + m.group(1)) if m else None

    def _detect_group_columns(self, rows):
        """Find the header row and its group columns. Returns
        (header_index, [(col_index, suffix), ...]); ([] if none found).
        The header is the first row (col 0 skipped) that names any group."""
        for idx, row in enumerate(rows):
            cols = [
                (c, self._group_suffix(cell))
                for c, cell in enumerate(row)
                if c >= 1 and self._group_suffix(cell)
            ]
            if cols:
                return idx, cols
        return None, []

    def _upsert_line(self, table_type, column_key, discount_code, pct, counters, group=None):
        """Create the discount line, or overwrite it if one with this
        (table_type, column_key, discount_code) already exists — whether from an
        earlier row in this sheet or a previous import (update-mode re-import)."""
        DiscountLine = self.env['baf.discount.line']
        existing = DiscountLine.search([
            ('table_type', '=', table_type),
            ('column_key', '=', column_key),
            ('discount_code', '=', discount_code),
        ], limit=1)
        if existing:
            vals = {}
            if existing.discount_pct != pct:
                vals['discount_pct'] = pct
            if group and existing.group_id != group:
                vals['group_id'] = group.id
            if vals:
                existing.write(vals)
            counters['updated'] += 1
        else:
            DiscountLine.create({
                'table_type': table_type,
                'column_key': column_key,
                'discount_code': discount_code,
                'discount_pct': pct,
                'group_id': group.id if group else False,
            })
            counters['created'] += 1

    def _ensure_group(self, name, suffix, counters):
        """Idempotently create a baf.sales.group with table_lookup pricing,
        scoped to the selected family so the pricing engine can pick it by the
        product's family when a customer holds several."""
        Group = self.env['baf.sales.group']
        existing = Group.search([('name', '=', name)], limit=1)
        if existing:
            vals = {}
            if existing.pricing_method != 'table_lookup':
                vals['pricing_method'] = 'table_lookup'
            if existing.group_column_suffix != suffix:
                vals['group_column_suffix'] = suffix
            if existing.family_id != self.family_id:
                vals['family_id'] = self.family_id.id
            if vals:
                existing.write(vals)
            return existing
        group = Group.create({
            'name': name,
            'pricing_method': 'table_lookup',
            'group_column_suffix': suffix,
            'family_id': self.family_id.id,
        })
        counters['groups_created'] += 1
        return group

    # ────────────────────────────────────────────────────────────────
    # Importer (variable group count; layout from the import method)
    # ────────────────────────────────────────────────────────────────

    def _check_family(self):
        if not self.family_id:
            raise UserError(_("Select the brand family this sheet prices."))
        if not self._brand_bases():
            raise UserError(_(
                "Family '%s' has no brands, so there is nothing to price. "
                "Add brands to it first.") % self.family_id.name)

    def _column_keys(self):
        """Discount column keys one group occupies, left to right. Types & Groups
        method: one key per (brand base, type bucket) — BMW_T12, BMW_T39,
        MINI_T12, MINI_T39. Groups Table method: one key per brand base."""
        bases = self._brand_bases()
        if self._is_type_split():
            return [f"{base}_{bucket}" for base in bases for bucket in BAF_TYPE_BUCKETS]
        return bases

    def _import_matrix(self, rows):
        counters = {'created': 0, 'updated': 0, 'removed': 0, 'groups_created': 0}

        self._check_family()
        column_keys = self._column_keys()
        type_split = self._is_type_split()

        header_idx, group_cols = self._detect_group_columns(rows)
        if not group_cols:
            raise UserError(_(
                "No group columns found. The header must name each group, "
                "e.g. 'SALE PRICE GR1', 'GR8', 'GR_MOTORCYCLE'."))

        # Replace mode wipes every existing sales line of this family first, so
        # the import that follows is the family's whole pricing. Update mode
        # keeps them and lets _upsert_line overwrite/add per code. Either way the
        # group records (and customer assignments) are kept.
        if self.write_mode == 'replace':
            self._clear_existing_lines(counters)

        # One sales group per detected section, all scoped to the family
        # (e.g. BMW_MINI_GR1 covers BMW_T12_GR1 ... MINI_T39_GR1).
        group_prefix = '_'.join(self._brand_bases())
        groups = {
            suffix: self._ensure_group(f"{group_prefix}_{suffix}", suffix, counters)
            for _base_col, suffix in group_cols
        }

        # Types & Groups spreads its keys across consecutive columns; Groups
        # Table feeds every brand base from the group's single column.
        for i, row in enumerate(rows):
            if i == header_idx or not row:
                continue
            code = str(row[0]).strip()
            if not code or code.upper() == 'DC':
                continue
            for base_col, suffix in group_cols:
                for offset, key in enumerate(column_keys):
                    col_idx = base_col + (offset if type_split else 0)
                    if col_idx >= len(row):
                        continue
                    pct = self._parse_float(row[col_idx])
                    if pct is None:
                        continue
                    self._upsert_line('sales', f"{key}_{suffix}", code, pct,
                                      counters, group=groups[suffix])

        return counters

    def _clear_existing_lines(self, counters):
        """Delete the sales discount lines of every group scoped to the selected
        family. Runs only in 'Delete existing groups' mode, so the import that
        follows starts from a clean slate. Group records are kept."""
        groups = self.env['baf.sales.group'].search([('family_id', '=', self.family_id.id)])
        lines = groups.mapped('discount_line_ids').filtered(
            lambda l: l.table_type == 'sales')
        counters['removed'] += len(lines)
        lines.unlink()

