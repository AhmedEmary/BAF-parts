import gzip
import io
import csv
from datetime import date
from urllib.parse import urlparse

from odoo.tests import HttpCase, tagged

from odoo.addons.b2b_custom.controllers.pricefile import pricefile_query, visible_brands


BMW_MINI_COLUMNS = [
    'Brand', 'Type', 'SKU', 'Replaced By', 'Description',
    'UPE', 'Discount Code', 'Surcharge', 'Unit of Sale', 'MOD',
]

JLR_COLUMNS = [
    'Brand', 'SKU', 'Replaced By', 'Discount Code', 'UPE',
    'Description', 'Surcharge', 'Unit of Sale',
]


@tagged('post_install', '-at_install')
class TestPriceFile(HttpCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        Family = cls.env['baf.brand.family']
        cls.fam_bmw = Family.create({'name': 'QA BMW/MINI'})
        cls.fam_jlr = Family.create({'name': 'QA JLR'})

        Brand = cls.env['product.brand']
        cls.brand_public = Brand.create({
            'name': 'QA-BMW', 'is_public': True, 'family_id': cls.fam_bmw.id})
        cls.brand_company = Brand.create({
            'name': 'QA-Jaguar', 'is_public': False, 'family_id': cls.fam_jlr.id})
        cls.brand_child = Brand.create({'name': 'QA-Mercedes-Benz', 'is_public': False})
        cls.brand_hidden = Brand.create({'name': 'QA-Bosal', 'is_public': False})

        cls.company_partner = cls.env['res.partner'].create({
            'name': 'B2B Company',
            'is_company': True,
            'visible_brand_ids': [(6, 0, cls.brand_company.ids)],
        })
        cls.child_partner = cls.env['res.partner'].create({
            'name': 'B2B Child Contact',
            'parent_id': cls.company_partner.id,
            'visible_brand_ids': [(6, 0, cls.brand_child.ids)],
        })

        Template = cls.env['product.template']
        cls.bmw_replacement = Template.create({
            'name': 'BMW Brake Pad V2',
            'sku': 'BMW-001-R',
            'brand': cls.brand_public.id,
            'list_price': 110.0,
        })
        cls.bmw_car = Template.create({
            'name': 'BMW Brake Pad',
            'sku': 'BMW-001',
            'brand': cls.brand_public.id,
            'list_price': 100.0,
            'surcharge': 5.0,
            'baf_discount_code': 'TESTQA1',
            'baf_type_code': 1,
            'baf_mod': 'car',
            'unit_of_sales': 2,
            'replaced_by_id': cls.bmw_replacement.id,
        })
        cls.jlr_part = Template.create({
            'name': 'Jaguar Filter',
            'sku': 'JAG-001',
            'brand': cls.brand_company.id,
            'list_price': 80.0,
            'surcharge': 2.0,
            'baf_discount_code': '1A',
            'unit_of_sales': 1,
        })

        cls.user = cls.env['res.users'].create({
            'name': 'B2B Portal User',
            'login': 'pricefile_user',
            'password': 'pricefile_user',
            'partner_id': cls.company_partner.id,
            'group_ids': [(6, 0, [
                cls.env.ref('base.group_portal').id,
                cls.env.ref('b2b_custom.group_b2b_customer').id,
            ])],
        })

    @property
    def _fixture_brand_ids(self):
        return {
            self.brand_public.id, self.brand_company.id,
            self.brand_child.id, self.brand_hidden.id,
        }

    def test_visible_brands_parent_takes_own_brands_only(self):
        brands = visible_brands(self.env, self.company_partner)
        self.assertEqual(
            set(brands.ids) & self._fixture_brand_ids,
            {self.brand_public.id, self.brand_company.id},
        )

    def test_visible_brands_child_unions_parent_brands(self):
        brands = visible_brands(self.env, self.child_partner)
        self.assertEqual(
            set(brands.ids) & self._fixture_brand_ids,
            {self.brand_public.id, self.brand_company.id, self.brand_child.id},
        )

    def test_visible_brands_excludes_unrelated_private_brand(self):
        brands = visible_brands(self.env, self.child_partner)
        self.assertNotIn(self.brand_hidden.id, brands.ids)

    def _rows(self, partner, brand):
        self.env.flush_all()
        sql, params = pricefile_query(partner, brand, 'en_US')
        self.env.cr.execute(sql, params)
        cols = [d.name for d in self.env.cr.description]
        return cols, [dict(zip(cols, r)) for r in self.env.cr.fetchall()]

    def test_bmw_mini_layout_has_expected_columns(self):
        cols, rows = self._rows(self.company_partner, self.brand_public)
        self.assertEqual(cols, BMW_MINI_COLUMNS)
        by_sku = {r['SKU']: r for r in rows}
        row = by_sku['BMW-001']
        self.assertEqual(row['Brand'], self.brand_public.name)
        self.assertEqual(row['Type'], 1)
        self.assertEqual(row['Replaced By'], 'BMW-001-R')
        self.assertEqual(row['Description'], 'BMW Brake Pad')
        self.assertEqual(float(row['UPE']), 100.0)
        self.assertEqual(row['Discount Code'], 'TESTQA1')
        self.assertEqual(float(row['Surcharge']), 5.0)
        self.assertEqual(row['Unit of Sale'], 2)
        self.assertEqual(row['MOD'], 'car')

    def test_jlr_layout_has_expected_columns(self):
        cols, rows = self._rows(self.company_partner, self.brand_company)
        self.assertEqual(cols, JLR_COLUMNS)
        by_sku = {r['SKU']: r for r in rows}
        row = by_sku['JAG-001']
        self.assertEqual(row['Brand'], self.brand_company.name)
        self.assertEqual(row['Replaced By'], '')
        self.assertEqual(row['Discount Code'], '1A')
        self.assertEqual(float(row['UPE']), 80.0)
        self.assertEqual(row['Description'], 'Jaguar Filter')
        self.assertEqual(float(row['Surcharge']), 2.0)
        self.assertEqual(row['Unit of Sale'], 1)

    def test_other_family_uses_jlr_layout(self):
        merc_family = self.env['baf.brand.family'].create({'name': 'QA Mercedes'})
        merc_brand = self.env['product.brand'].create({
            'name': 'QA-Merc', 'is_public': True, 'family_id': merc_family.id})
        self.env['product.template'].create({
            'name': 'Merc Filter', 'sku': 'MRC-1', 'brand': merc_brand.id,
            'list_price': 50.0,
        })
        cols, _rows = self._rows(self.company_partner, merc_brand)
        self.assertEqual(cols, JLR_COLUMNS)

    def test_brand_column_shows_display_label_when_set(self):
        self.brand_public.display_label = 'Bayerische Motoren Werke'
        self.brand_company.display_label = 'Jaguar & Land Rover'
        _cols, bmw_rows = self._rows(self.company_partner, self.brand_public)
        _cols, jag_rows = self._rows(self.company_partner, self.brand_company)
        self.assertEqual(bmw_rows[0]['Brand'], 'Bayerische Motoren Werke')
        self.assertEqual(jag_rows[0]['Brand'], 'Jaguar & Land Rover')

    def test_brand_column_falls_back_to_technical_name_when_label_missing(self):
        self.brand_public.display_label = False
        _cols, rows = self._rows(self.company_partner, self.brand_public)
        self.assertEqual(rows[0]['Brand'], self.brand_public.name)

    def test_query_only_returns_the_requested_brand(self):
        _cols, rows = self._rows(self.company_partner, self.brand_public)
        skus = {r['SKU'] for r in rows}
        self.assertNotIn('JAG-001', skus)

    def test_inactive_and_non_sellable_products_excluded(self):
        Template = self.env['product.template']
        Template.create({
            'name': 'BMW Inactive', 'sku': 'BMW-INACTIVE',
            'brand': self.brand_public.id, 'list_price': 10.0, 'active': False,
        })
        Template.create({
            'name': 'BMW Not Sellable', 'sku': 'BMW-NOSELL',
            'brand': self.brand_public.id, 'list_price': 10.0, 'sale_ok': False,
        })
        _cols, rows = self._rows(self.company_partner, self.brand_public)
        skus = {r['SKU'] for r in rows}
        self.assertNotIn('BMW-INACTIVE', skus)
        self.assertNotIn('BMW-NOSELL', skus)
        self.assertIn('BMW-001', skus)

    def _download(self, brand_id):
        self.authenticate('pricefile_user', 'pricefile_user')
        return self.url_open(
            '/pricefile/download?brand_id=%s' % brand_id, allow_redirects=False
        )

    def _csv_rows(self, response):
        raw = response.content.decode('utf-8-sig')
        return list(csv.DictReader(io.StringIO(raw)))

    def test_download_returns_csv_for_visible_brand(self):
        response = self._download(self.brand_public.id)
        self.assertEqual(response.status_code, 200)
        self.assertIn('text/csv', response.headers['Content-Type'])
        self.assertEqual(response.headers['Content-Encoding'], 'gzip')

        expected_name = 'PriceList_%s_%s.csv' % (
            self.brand_public.name, date.today().isoformat())
        self.assertIn(expected_name, response.headers['Content-Disposition'])

        content = response.content
        if content[:2] == b'\x1f\x8b':
            content = gzip.decompress(content)
        self.assertTrue(content.startswith(b'\xef\xbb\xbf'), 'missing utf-8-sig BOM')

        rows = self._csv_rows(response)
        self.assertEqual(list(rows[0].keys()), BMW_MINI_COLUMNS)
        by_sku = {r['SKU']: r for r in rows}
        self.assertEqual(by_sku['BMW-001']['Description'], 'BMW Brake Pad')
        self.assertEqual(float(by_sku['BMW-001']['UPE']), 100.0)
        self.assertEqual(by_sku['BMW-001']['Discount Code'], 'TESTQA1')
        self.assertEqual(by_sku['BMW-001']['Replaced By'], 'BMW-001-R')

    def test_download_non_latin1_brand_name(self):
        brand = self.env['product.brand'].create({'name': 'Škoda', 'is_public': True})
        self.env['product.template'].create({
            'name': 'Skoda Wiper Blade', 'sku': 'SKO-001',
            'brand': brand.id, 'list_price': 20.0,
        })
        response = self._download(brand.id)
        self.assertEqual(response.status_code, 200)

    def _assert_redirects_to_pricefile(self, response):
        self.assertEqual(response.status_code, 303)
        location = response.headers['Location']
        self.assertTrue(urlparse(location).path.endswith('/pricefile'), location)
        self.assertNotIn('brand_id', location)

    def test_download_rejects_brand_the_partner_cannot_see(self):
        response = self._download(self.brand_hidden.id)
        self._assert_redirects_to_pricefile(response)

    def test_download_rejects_garbage_brand_id(self):
        response = self._download('not-a-number')
        self._assert_redirects_to_pricefile(response)

    def test_download_rejects_missing_brand_id(self):
        self.authenticate('pricefile_user', 'pricefile_user')
        response = self.url_open('/pricefile/download', allow_redirects=False)
        self._assert_redirects_to_pricefile(response)

    def test_download_rejects_negative_brand_id(self):
        response = self._download(-1)
        self._assert_redirects_to_pricefile(response)

    def test_pricefile_page_renders_form_and_visible_brands(self):
        self.authenticate('pricefile_user', 'pricefile_user')
        response = self.url_open('/pricefile')
        self.assertEqual(response.status_code, 200)
        body = response.content.decode('utf-8')
        self.assertIn('action="/pricefile/download"', body)
        self.assertIn('name="brand_id"', body)
        self.assertIn(
            '<option value="%d">%s</option>' % (self.brand_company.id, self.brand_company.name),
            body,
        )
        self.assertNotIn('value="%d"' % self.brand_hidden.id, body)

    def _open_discount_workbook(self):
        import openpyxl  # noqa: I001
        self.authenticate('pricefile_user', 'pricefile_user')
        response = self.url_open(
            '/pricefile/discount-download', allow_redirects=False)
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'spreadsheetml.sheet', response.headers['Content-Type'])
        return response, openpyxl.load_workbook(
            io.BytesIO(response.content), read_only=False, data_only=True)

    def test_discount_download_one_sheet_per_visible_brand_with_rates(self):
        DiscountLine = self.env['baf.discount.line']
        group_bmw_gr1 = self.env['baf.sales.group'].create({
            'name': 'BMW GR1', 'family_id': self.fam_bmw.id,
            'pricing_method': 'table_lookup', 'group_column_suffix': 'GR1'})
        group_jlr_markup = self.env['baf.sales.group'].create({
            'name': 'JLR markup', 'family_id': self.fam_jlr.id,
            'pricing_method': 'markup_pct', 'markup_pct': 15.0})
        DiscountLine.create({
            'table_type': 'sales',
            'column_key': '%s_GR1' % self.bmw_car.baf_sales_column_key,
            'discount_code': 'TESTQA1', 'discount_pct': 20.0,
        })
        self.company_partner.sales_group_ids = [
            (6, 0, [group_bmw_gr1.id, group_jlr_markup.id])]
        response, wb = self._open_discount_workbook()
        self.assertEqual(wb.sheetnames, [self.brand_public.display_name])
        expected_name = 'DiscountCodes_%s_%s.xlsx' % (
            self.company_partner.name.replace(' ', '_'), date.today().isoformat())
        self.assertIn(expected_name, response.headers['Content-Disposition'])

    def test_discount_download_pivots_rate_per_type_bucket(self):
        group_bmw_gr1 = self.env['baf.sales.group'].create({
            'name': 'BMW GR1', 'family_id': self.fam_bmw.id,
            'pricing_method': 'table_lookup', 'group_column_suffix': 'GR1'})
        self.env['baf.discount.line'].create({
            'table_type': 'sales',
            'column_key': '%s_GR1' % self.bmw_car.baf_sales_column_key,
            'discount_code': 'TESTQA1', 'discount_pct': 20.0,
        })
        self.company_partner.sales_group_ids = [(6, 0, [group_bmw_gr1.id])]
        _response, wb = self._open_discount_workbook()
        ws = wb[self.brand_public.display_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(header[0], 'Discount Code')
        self.assertIn('TA 1-2-4-6-8', header)
        t12_col = header.index('TA 1-2-4-6-8')
        by_code = {
            row[0].value: row[t12_col].value
            for row in ws.iter_rows(min_row=2)
        }
        self.assertEqual(by_code['TESTQA1'], 20.0)

    def test_discount_download_redirects_when_no_applicable_rates(self):
        group_jlr_markup = self.env['baf.sales.group'].create({
            'name': 'JLR markup', 'family_id': self.fam_jlr.id,
            'pricing_method': 'markup_pct', 'markup_pct': 15.0})
        self.company_partner.sales_group_ids = [(6, 0, [group_jlr_markup.id])]
        self.authenticate('pricefile_user', 'pricefile_user')
        response = self.url_open(
            '/pricefile/discount-download', allow_redirects=False)
        self.assertEqual(response.status_code, 303)
        self.assertTrue(
            urlparse(response.headers['Location']).path.endswith('/pricefile'))

    def test_pricefile_page_advertises_discount_download(self):
        self.authenticate('pricefile_user', 'pricefile_user')
        response = self.url_open('/pricefile')
        self.assertEqual(response.status_code, 200)
        self.assertIn('/pricefile/discount-download', response.content.decode('utf-8'))
