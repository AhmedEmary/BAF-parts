import csv
import io
import logging
import re

from odoo import http
from odoo.http import request

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

_logger = logging.getLogger(__name__)


def _normalize_sku(value):
    if value is None:
        return ''
    return re.sub(r'\s+', '', str(value)).upper()


def _format_price(template, partner):
    try:
        price = template.with_context(partner=partner).baf_website_display_price()
    except Exception:
        price = template.list_price or 0.0
    return _format_amount(template, price)


def _format_amount(template, amount):
    currency = template.currency_id or request.env.company.currency_id
    symbol = currency.symbol or '€'
    return f"{amount or 0.0:,.2f} {symbol}".replace(',', 'X').replace('.', ',').replace('X', '.')


def _visible_brand_domain(partner):
    """Restrict product.product search to brands the partner may see.

    Rules (mirrors `website.sale_product_domain`):
    - Public brands and no-brand products are always visible.
    - If the partner's commercial (company) partner has `visible_brand_ids`
      set, those brands are visible on top.
    `visible_brand_ids` is read off `commercial_partner_id` so a logged-in
    employee inherits the company's brand access.
    """
    commercial = partner.commercial_partner_id if partner else partner
    visible_ids = commercial.visible_brand_ids.ids if commercial else []
    if visible_ids:
        return [
            '|', '|',
            ('product_tmpl_id.brand.is_public', '=', True),
            ('product_tmpl_id.brand', '=', False),
            ('product_tmpl_id.brand', 'in', visible_ids),
        ]
    return [
        '|',
        ('product_tmpl_id.brand.is_public', '=', True),
        ('product_tmpl_id.brand', '=', False),
    ]


def _mod_label(template):
    value = template.baf_mod
    if not value:
        return ''
    selection = dict(template._fields['baf_mod'].selection)
    return selection.get(value, value)


def _type_label(template):
    code = template.baf_type_code or 0
    return str(code) if code else ''


def _product_to_dict(product, partner):
    template = product.product_tmpl_id
    is_nla = template._baf_is_nla()
    blocked = template._baf_is_order_blocked()
    qty = product.free_qty if hasattr(product, 'free_qty') else product.qty_available

    if is_nla:
        availability = 'NLA'
        availability_type = 'nla'
        delivery_time = ''
    elif blocked and template.replaced_by_id:
        availability = 'Ersetzt – Nachfolger bestellen'
        availability_type = 'replacement'
        delivery_time = ''
    elif blocked:
        availability = 'Nicht mehr verfügbar'
        availability_type = 'no'
        delivery_time = ''
    elif qty > 5:
        availability = 'Sofort verfügbar'
        availability_type = 'ok'
        delivery_time = '1-2 Werktage'
    elif qty > 0:
        availability = 'Geringer Bestand'
        availability_type = 'low'
        delivery_time = '2-4 Werktage'
    else:
        availability = ''
        availability_type = ''
        delivery_time = ''

    moq = int(template.unit_of_sales or 0) or 1
    surcharge_raw = getattr(product, 'surcharge', 0.0) or getattr(template, 'surcharge', 0.0) or 0.0
    replacement = template.replaced_by_id

    return {
        'id': product.id,
        'part_number': template.sku or product.default_code or '',
        'name': template.name,
        'brand': template.brand.name if template.brand else '',
        'brand_id': template.brand.id if template.brand else 0,
        'mod': _mod_label(template),
        'type': _type_label(template),
        'price': _format_price(template, partner),
        'moq': moq,
        'surcharge': _format_amount(template, surcharge_raw) if surcharge_raw else '',
        'surcharge_raw': surcharge_raw,
        'availability': availability,
        'availability_type': availability_type,
        'delivery_time': delivery_time,
        'orderable': not blocked,
        'is_nla': is_nla,
        'replacement_url': (
            replacement.website_url if blocked and replacement else ''
        ),
        'replacement_name': (
            replacement.display_name if blocked and replacement else ''
        ),
        'replacement_part_number': (
            replacement.sku or '' if blocked and replacement else ''
        ),
        'replacement_product_id': (
            replacement.product_variant_id.id if blocked and replacement else 0
        ),
    }


class BafB2BController(http.Controller):

    @http.route(['/bestellsystem'], type='http', auth='user', website=True, sitemap=False)
    def baf_b2b_page(self, **kwargs):
        return request.render('b2b_custom.baf_b2b_page', {})

    @http.route(
        '/bestellsystem/part-search', type='jsonrpc', auth='user', methods=['POST'], csrf=False, website=True,
    )
    def part_search(self, part_numbers=None, brand_choices=None, quantities=None,
                    product_ids=None, **kwargs):
        """Search products by SKU.

        Returns:
          products:  list of product dicts (unique SKU match, or brand-resolved)
          ambiguous: list of {part_number, matches:[{product_id, brand, brand_id, name}, ...]}
          not_found: list of part numbers with no match
        """
        part_numbers = list(part_numbers or [])
        product_ids = [int(pid) for pid in (product_ids or []) if pid]
        if not part_numbers and not product_ids:
            return {'products': [], 'ambiguous': [], 'not_found': []}

        partner = request.env.user.partner_id
        Product = request.env['product.product'].sudo()

        # Normalize quantity map for echo back to client
        qty_map = {}
        for key, value in (quantities or {}).items():
            try:
                qty_map[_normalize_sku(key)] = max(1, int(float(value)))
            except (TypeError, ValueError):
                continue

        # Direct lookup by product id (used by "Nachfolger ansehen" flow)
        direct_products = []
        if product_ids:
            for product in Product.browse(product_ids).exists():
                if not product.product_tmpl_id.active or not product.product_tmpl_id.sale_ok:
                    continue
                direct_products.append(product)

        # 1. Normalize + dedupe input
        normalized = []
        seen = set()
        for raw in part_numbers:
            key = _normalize_sku(raw)
            if key and key not in seen:
                seen.add(key)
                normalized.append({'raw': str(raw).strip(), 'key': key})

        # 2. Fetch every matching variant in a single query (case-insensitive),
        #    restricted to brands this partner is allowed to see.
        keys = [item['key'] for item in normalized]
        templates_by_key = {}
        if keys:
            products = Product.search([
                ('sku', 'in', keys + [k.lower() for k in keys]),
                ('active', '=', True),
                ('sale_ok', '=', True),
            ] + _visible_brand_domain(partner))
            for product in products:
                key = _normalize_sku(product.product_tmpl_id.sku)
                templates_by_key[key] = templates_by_key.get(key, Product.browse([])) | product

        # 3. brand_choices = {normalized_sku: brand_id} — resolves a previous ambiguity
        brand_choices = brand_choices or {}
        brand_choices = {_normalize_sku(k): int(v) for k, v in brand_choices.items() if v}

        products_out = []
        ambiguous = []
        not_found = []

        for item in normalized:
            matches = templates_by_key.get(item['key'])
            if not matches:
                not_found.append(item['raw'])
                continue

            # Group variants by brand; if user pre-picked a brand, filter to it
            picked_brand = brand_choices.get(item['key'])
            if picked_brand:
                matches = matches.filtered(
                    lambda p: p.product_tmpl_id.brand.id == picked_brand
                )
                if not matches:
                    not_found.append(item['raw'])
                    continue

            brand_groups = {}
            for product in matches:
                brand_id = product.product_tmpl_id.brand.id if product.product_tmpl_id.brand else 0
                brand_groups[brand_id] = brand_groups.get(brand_id, Product.browse([])) | product

            if len(brand_groups) > 1:
                ambiguous.append({
                    'part_number': item['raw'],
                    'matches': [
                        {
                            'product_id': group[:1].id,
                            'brand_id': group[:1].product_tmpl_id.brand.id if group[:1].product_tmpl_id.brand else 0,
                            'brand': group[:1].product_tmpl_id.brand.name if group[:1].product_tmpl_id.brand else '—',
                            'name': group[:1].product_tmpl_id.name,
                        }
                        for group in brand_groups.values() if group
                    ],
                })
                continue

            # Unambiguous: pick the first variant of the single matching brand
            chosen = next(iter(brand_groups.values()))[:1]
            data = _product_to_dict(chosen, partner)
            requested_qty = qty_map.get(item['key'])
            if requested_qty:
                data['requested_quantity'] = requested_qty
            products_out.append(data)

        # Append direct product-id lookups (from "Nachfolger ansehen") at the
        # end. Skip duplicates that were already resolved via SKU above.
        already_in = {p['id'] for p in products_out}
        for product in direct_products:
            if product.id in already_in:
                continue
            data = _product_to_dict(product, partner)
            qty_key = _normalize_sku(data.get('part_number'))
            requested_qty = qty_map.get(qty_key)
            if requested_qty:
                data['requested_quantity'] = requested_qty
            products_out.append(data)

        echo_quantities = {sku: qty_map[sku] for sku in qty_map}

        return {
            'products': products_out,
            'ambiguous': ambiguous,
            'not_found': not_found,
            'quantities': echo_quantities,
        }

    @http.route(
        '/bestellsystem/cart/add', type='jsonrpc', auth='user', methods=['POST'], csrf=False, website=True,
    )
    def cart_add(self, items=None, **kwargs):
        if not items:
            return {'success': False, 'message': 'Keine Artikel übergeben.'}

        order = request.cart or request.website._create_cart()
        Product = request.env['product.product'].sudo()

        added = 0
        failed = []
        for item in items:
            try:
                product_id = int(item.get('product_id'))
                qty = float(item.get('quantity') or 1)
            except (TypeError, ValueError):
                failed.append(item)
                continue
            if qty <= 0:
                continue
            product = Product.browse(product_id).exists()
            if not product:
                failed.append({'product_id': product_id, 'error': 'not_found'})
                continue
            if product.product_tmpl_id._baf_is_order_blocked():
                failed.append({
                    'product_id': product_id,
                    'error': 'not_orderable',
                    'sku': product.product_tmpl_id.sku or product.default_code or '',
                })
                continue
            note = (item.get('note') or '').strip()
            try:
                result = order.with_context(skip_cart_verification=True)._cart_add(
                    product_id=product.id, quantity=qty,
                )
                if note:
                    line_id = (result or {}).get('line_id')
                    line = (
                        request.env['sale.order.line'].sudo().browse(line_id)
                        if line_id
                        else order.order_line.filtered(
                            lambda l: l.product_id.id == product.id
                        )[-1:]
                    )
                    if line:
                        line.sudo().write({'baf_line_note': note})
                added += 1
            except Exception as exc:
                _logger.exception("B2B cart_add failed for product %s: %s", product_id, exc)
                failed.append({'product_id': product_id, 'error': str(exc)})

        cart_qty = int(order.cart_quantity) if order else 0
        return {
            'success': added > 0,
            'added': added,
            'failed': failed,
            'cart_quantity': cart_qty,
            'cart_url': '/shop/cart',
        }

    @http.route(
        '/bestellsystem/upload-parts-list', type='http', auth='user', methods=['POST'], csrf=False, website=True,
    )
    def upload_parts_list(self, **post):
        upload = post.get('file')
        if not upload:
            return request.make_json_response(
                {'products': [], 'ambiguous': [], 'not_found': [], 'error': 'Keine Datei erhalten.'},
                status=400,
            )

        try:
            part_numbers = self._extract_part_numbers(upload)
        except Exception as exc:
            _logger.exception("B2B upload parse failed: %s", exc)
            return request.make_json_response(
                {'products': [], 'ambiguous': [], 'not_found': [],
                 'error': 'Datei konnte nicht gelesen werden.'},
                status=400,
            )

        result = self.part_search(part_numbers=part_numbers)
        return request.make_json_response(result)

    # ── helpers ───────────────────────────────────────────────────────────────

    def _extract_part_numbers(self, upload):
        filename = (upload.filename or '').lower()
        data = upload.read()
        if not data:
            return []

        if filename.endswith('.csv') or filename.endswith('.txt'):
            return self._parse_csv_bytes(data)
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            return self._parse_xlsx_bytes(data)
        # Best-effort fallback: treat as CSV
        return self._parse_csv_bytes(data)

    def _parse_csv_bytes(self, data):
        for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                text = data.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = data.decode('utf-8', errors='ignore')

        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(io.StringIO(text), dialect)

        rows = list(reader)
        if not rows:
            return []
        header = [(cell or '').strip().lower() for cell in rows[0]]
        sku_col = self._find_sku_column(header)
        start_row = 1 if sku_col is not None else 0
        if sku_col is None:
            sku_col = 0

        result = []
        for row in rows[start_row:]:
            if len(row) <= sku_col:
                continue
            value = (row[sku_col] or '').strip()
            if value:
                result.append(value)
        return result

    def _parse_xlsx_bytes(self, data):
        if load_workbook is None:
            raise RuntimeError('openpyxl ist nicht installiert.')
        workbook = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        header = [(str(cell or '')).strip().lower() for cell in rows[0]]
        sku_col = self._find_sku_column(header)
        start_row = 1 if sku_col is not None else 0
        if sku_col is None:
            sku_col = 0

        result = []
        for row in rows[start_row:]:
            if not row or len(row) <= sku_col:
                continue
            value = row[sku_col]
            if value is None:
                continue
            value = str(value).strip()
            if value:
                result.append(value)
        return result

    def _find_sku_column(self, header):
        candidates = ('teilenummer', 'sku', 'part_number', 'part number',
                      'artikel', 'artikelnr', 'code', 'reference', 'referenz')
        for idx, cell in enumerate(header):
            if cell in candidates:
                return idx
        return None