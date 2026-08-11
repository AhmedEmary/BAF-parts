import gzip
import io
import re
from datetime import date

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

from odoo import http, _
from odoo.addons.general_system_custom.models.baf_product_pricing import (
    baf_brand_base_key,
    baf_family_base_key,
    BAF_TYPE_SPLIT_FAMILIES,
    resolve_baf_brand_info,
)
from odoo.addons.portal.controllers.portal import CustomerPortal
from odoo.exceptions import UserError
from odoo.http import content_disposition, request


def visible_brands(env, partner):
    """Brands the partner may pull a price file for: publicly available brands,
    plus his own visible brands, plus — for a child contact — the brands visible
    to his company."""
    brand_ids = set(partner.visible_brand_ids.ids)
    commercial = partner.commercial_partner_id
    if commercial and commercial != partner:
        brand_ids |= set(commercial.visible_brand_ids.ids)
    return env['product.brand'].sudo().search(
        ['|', ('is_public', '=', True), ('id', 'in', list(brand_ids))],
        order='name',
    )


# BMW/MINI export columns (task 1.2). Raw UPE + Discount Code — the customer
# applies the discount table themselves; the pricefile carries the ingredients.
_PRICEFILE_SQL_BMW_MINI = """
    SELECT
        COALESCE(pb.display_label->>%(lang)s, pb.display_label->>'en_US', pb.name, '') AS "Brand",
        pt.baf_type_code AS "Type",
        pt.sku AS "SKU",
        COALESCE(rpl.sku, '') AS "Replaced By",
        COALESCE(pt.name->>%(lang)s, pt.name->>'en_US') AS "Description",
        round(COALESCE(pt.list_price, 0)::numeric, 2) AS "UPE",
        COALESCE(pt.baf_discount_code, '') AS "Discount Code",
        round(COALESCE(pt.surcharge, 0)::numeric, 2) AS "Surcharge",
        pt.unit_of_sales AS "Unit of Sale",
        COALESCE(pt.baf_mod, '') AS "MOD"
    FROM product_template pt
    LEFT JOIN product_brand pb ON pb.id = pt.brand
    LEFT JOIN product_template rpl ON rpl.id = pt.replaced_by_id
    WHERE pt.active AND pt.sale_ok AND pt.brand = ANY(%(brand_ids)s)
"""

# JLR + everything else — same layout for all non-BMW/MINI families.
_PRICEFILE_SQL_JLR = """
    SELECT
        COALESCE(pb.display_label->>%(lang)s, pb.display_label->>'en_US', pb.name, '') AS "Brand",
        pt.sku AS "SKU",
        COALESCE(rpl.sku, '') AS "Replaced By",
        COALESCE(pt.baf_discount_code, '') AS "Discount Code",
        round(COALESCE(pt.list_price, 0)::numeric, 2) AS "UPE",
        COALESCE(pt.name->>%(lang)s, pt.name->>'en_US') AS "Description",
        round(COALESCE(pt.surcharge, 0)::numeric, 2) AS "Surcharge",
        pt.unit_of_sales AS "Unit of Sale"
    FROM product_template pt
    LEFT JOIN product_brand pb ON pb.id = pt.brand
    LEFT JOIN product_template rpl ON rpl.id = pt.replaced_by_id
    WHERE pt.active AND pt.sale_ok AND pt.brand = ANY(%(brand_ids)s)
"""


def pricefile_query(partner, brand, lang):
    return _pricefile_query_for_brands(brand, brand.family_id, lang)


def pricefile_query_for_family(partner, family, lang):
    brands = family.brand_ids if family else False
    if not brands:
        return None, None
    return _pricefile_query_for_brands(brands, family, lang)


def _pricefile_query_for_brands(brands, product_bfam, lang):
    label = product_bfam.name if product_bfam else (
        brands[0].name if len(brands) else '')
    family = resolve_baf_brand_info(label)[1]
    sql = _PRICEFILE_SQL_BMW_MINI if family == 'bmw_mini' else _PRICEFILE_SQL_JLR
    params = {'lang': lang, 'brand_ids': list(brands.ids)}
    return sql, params


def visible_families(env, partner):
    """Brand families the customer sees — a family is visible when at least
    one of its brands is."""
    brands = visible_brands(env, partner)
    families = brands.mapped('family_id').filtered(lambda f: f.active)
    return families.sorted('name')


# Type-bucket labels shared with the discount export. The bucket is the last
# `_T12` / `_T39` fragment of a discount-line column_key stripped of the
# group suffix. Only the BMW/MINI family has this split today; every other
# family has a single bucket keyed to the family base ('JLR', 'MERCEDES', ...).
_TYPE_BUCKET_LABELS = {
    'T12': "TA 1-2-4-6-8",
    'T39': "TA 3-5-7-9",
}
# Excel sheet-name reserved chars + 31-char cap. Trims the brand name to fit.
_SHEET_NAME_INVALID_RE = re.compile(r'[\\/*?:\[\]]')


def _safe_sheet_name(name, used):
    """Return an Excel-valid, unique sheet name derived from `name`."""
    cleaned = _SHEET_NAME_INVALID_RE.sub(' ', (name or 'Brand')).strip() or 'Brand'
    cleaned = cleaned[:31]
    base = cleaned
    suffix = 2
    while cleaned in used:
        # Reserve room for the ' (N)' suffix so the cap is respected.
        tag = ' (%d)' % suffix
        cleaned = (base[: 31 - len(tag)]) + tag
        suffix += 1
    used.add(cleaned)
    return cleaned


def _applicable_sales_groups(partner, brand):
    """Every active sales group the partner has for this brand's family: the
    family-scoped ones (car + optional moto), and the wildcard fallback when no
    family-scoped group covers the brand. Mirrors `pricefile_query`'s resolution
    so the exported %s can never disagree with the price on the pricefile."""
    groups = partner._baf_effective_sales_groups().filtered(lambda g: g.active)
    family = brand.family_id
    family_groups = groups.filtered(lambda g: family and g.family_id == family)
    if family_groups:
        return family_groups
    return groups.filtered(lambda g: not g.family_id)[:1]


def _column_label(bucket, is_moto):
    """User-facing header for a (type bucket, moto?) column. Never carries a
    group name or the group's column suffix — customers must not see internal
    grouping (task #49)."""
    parts = []
    if bucket in _TYPE_BUCKET_LABELS:
        parts.append(_TYPE_BUCKET_LABELS[bucket])
    if is_moto:
        parts.append(_("Motorcycle"))
    return " ".join(parts) if parts else _("Discount %")


def _brand_sales_key_bases(brand):
    """Every base that `baf_sales_column_key` can take for products of this
    brand, paired with its public type bucket. Mirrors
    `_compute_baf_column_key`: for type-split families (BMW/MINI) the base is
    the brand's own base and there is one per type bucket
    (e.g. 'QA_BMW_T12', 'QA_BMW_T39'); for every other family the base is the
    family base and there is a single, bucket-less entry ('JLR'). Returning
    the same bases the pricing engine writes lets us reconstruct the discount
    line column_keys exactly and search for them by equality."""
    brand_name = brand.name or ''
    family_kind = resolve_baf_brand_info(brand_name)[1]
    if family_kind in BAF_TYPE_SPLIT_FAMILIES:
        return [
            ('T12', resolve_baf_brand_info(brand_name, type_code=1)[0]),
            ('T39', resolve_baf_brand_info(brand_name, type_code=3)[0]),
        ]
    if brand.family_id:
        return [('', baf_family_base_key(brand.family_id))]
    return [('', baf_brand_base_key(brand_name))]


def _brand_discount_rows(brand, groups):
    """Return (headers, rows) for one brand's discount worksheet, or (None, None)
    when the brand has no applicable table-lookup rate for this customer.

    Row shape: `[discount_code, pct_col_1, pct_col_2, ...]` — one column per
    (type bucket, moto flag) combination that actually has data. Column
    headers are the public labels only; the group's column suffix (GR1 / GR2 /
    MOTO / etc.) never leaks into the sheet.

    Discount lines are keyed by (table_type, column_key, discount_code) with
    column_key = `<base>_<group_suffix>`. That's how the pricing engine looks
    them up (see pricefile_query / _PRICEFILE_SQL); the lines themselves don't
    necessarily carry a group_id, so we match on column_key values rebuilt
    from the customer's applicable groups and the brand's sales-key bases."""
    if not groups:
        return None, None
    lookup_groups = groups.filtered(lambda g: g.pricing_method == 'table_lookup')
    if not lookup_groups:
        return None, None
    bases = [(bucket, base) for bucket, base in _brand_sales_key_bases(brand) if base]
    if not bases:
        return None, None

    # {full_column_key: (bucket, is_moto)} — used to decorate lookup results.
    # 'GR1' matches pricefile_query's default when a group has no explicit
    # suffix, so column_keys stay in lockstep with the priced pricefile.
    key_meta = {}
    for group in lookup_groups:
        suffix = (group.group_column_suffix or 'GR1').strip()
        is_moto = group._is_moto_group()
        for bucket, base in bases:
            key = '%s_%s' % (base, suffix)
            key_meta[key] = (bucket, is_moto)
    if not key_meta:
        return None, None

    lines = request.env['baf.discount.line'].sudo().search([
        ('table_type', '=', 'sales'),
        ('column_key', 'in', list(key_meta.keys())),
    ])
    if not lines:
        return None, None

    # {(bucket, is_moto): {code: pct}} — later flattened to a table.
    cells = {}
    for line in lines:
        meta = key_meta.get(line.column_key)
        if not meta:
            continue
        cells.setdefault(meta, {})[line.discount_code] = line.discount_pct

    if not cells:
        return None, None
    # Consistent column order: car buckets first (T12, T39, ''), then moto.
    def _col_sort_key(col):
        bucket, is_moto = col
        bucket_order = {'T12': 0, 'T39': 1, '': 2}.get(bucket, 3)
        return (int(is_moto), bucket_order)
    columns = sorted(cells.keys(), key=_col_sort_key)
    all_codes = sorted({
        code for col in columns for code in cells[col].keys()
    })
    headers = [_("Discount Code")] + [_column_label(*col) for col in columns]
    rows = []
    for code in all_codes:
        row = [code]
        for col in columns:
            row.append(cells[col].get(code))
        rows.append(row)
    return headers, rows


class PriceFile(CustomerPortal):

    @http.route(['/pricefile'], type='http', auth='user', website=True)
    def pricefile_page(self, **kw):
        partner = request.env.user.partner_id
        brands = visible_brands(request.env, partner)
        families = visible_families(request.env, partner)
        etk = request.env['baf.etk.file']._baf_current()
        return request.render('b2b_custom.pricefile_page', {
            'brands': brands,
            'families': families,
            'etk_available': bool(etk),
        })

    @http.route(
        ['/pricefile/etk'], type='http', auth='user', website=True,
        multilang=False, sitemap=False,
    )
    def etk_download(self, **kw):
        import base64
        etk = request.env['baf.etk.file']._baf_current()
        if not etk or not etk.file_data:
            return request.redirect('/pricefile')
        data = base64.b64decode(etk.file_data)
        filename = etk.file_name or 'ETK.bin'
        return request.make_response(data, headers=[
            ('Content-Type', 'application/octet-stream'),
            ('Content-Length', str(len(data))),
            ('Content-Disposition', content_disposition(filename)),
        ])

    @http.route(
        ['/pricefile/discount-download'], type='http', auth='user', website=True,
        multilang=False, sitemap=False,
    )
    def discount_download(self, **kw):
        """One xlsx file for the customer, one worksheet per visible brand,
        listing the sales discount codes and effective % that apply to them.
        Sheets are omitted for brands whose applicable group is markup-only or
        has no rows, so the file only shows brands the customer actually gets
        table-lookup rates for. Internal group names / column suffixes never
        appear (task #49)."""
        if not openpyxl:
            raise UserError(_(
                "The 'openpyxl' Python library is required to export discount "
                "tables. Please contact your administrator."))
        partner = request.env.user.partner_id
        brands = visible_brands(request.env, partner)
        if not brands:
            return request.redirect('/pricefile')

        wb = openpyxl.Workbook()
        # Wipe the default empty sheet: we add per-brand sheets and don't want a
        # stray "Sheet" left over when we're done.
        wb.remove(wb.active)
        used_names = set()
        for brand in brands:
            groups = _applicable_sales_groups(partner, brand)
            headers, rows = _brand_discount_rows(brand, groups)
            if not headers:
                continue
            ws = wb.create_sheet(title=_safe_sheet_name(brand.display_name, used_names))
            ws.append(headers)
            for row in rows:
                ws.append(row)

        if not wb.sheetnames:
            # Every brand fell out (markup-only, no lines) — no file to serve.
            return request.redirect('/pricefile')

        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()
        buf.close()

        filename = "DiscountCodes_%s_%s.xlsx" % (
            (partner.name or 'Customer').replace(' ', '_'),
            date.today().isoformat(),
        )
        return request.make_response(data, headers=[
            ('Content-Type',
             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('Content-Length', str(len(data))),
            ('Content-Disposition', content_disposition(filename)),
        ])

    @http.route(
        ['/pricefile/download'], type='http', auth='user', website=True,
        multilang=False, sitemap=False,
    )
    def pricefile_download(self, brand_id=None, **kw):
        partner = request.env.user.partner_id
        try:
            brand = visible_brands(request.env, partner).filtered(
                lambda b: b.id == int(brand_id)
            )
        except (TypeError, ValueError):
            brand = None
        if not brand:
            return request.redirect('/pricefile')

        lang = request.env.context.get('lang') or 'en_US'
        sql, params = pricefile_query(partner, brand, lang)
        return self._pricefile_csv_response(
            sql, params,
            filename="PriceList_%s_%s.csv" % (
                brand.display_name or 'Brand', date.today().isoformat()),
        )

    @http.route(
        ['/pricefile/family/download'], type='http', auth='user', website=True,
        multilang=False, sitemap=False,
    )
    def pricefile_family_download(self, family_id=None, **kw):
        partner = request.env.user.partner_id
        try:
            fid = int(family_id)
        except (TypeError, ValueError):
            return request.redirect('/pricefile')
        family = visible_families(request.env, partner).filtered(
            lambda f: f.id == fid)
        if not family:
            return request.redirect('/pricefile')

        lang = request.env.context.get('lang') or 'en_US'
        sql, params = pricefile_query_for_family(partner, family, lang)
        if not sql:
            return request.redirect('/pricefile')
        filename = "PriceList_%s_%s.csv" % (
            (family.name or 'BrandGroup').replace(' ', '_'),
            date.today().isoformat(),
        )
        return self._pricefile_csv_response(sql, params, filename=filename)

    def _pricefile_csv_response(self, sql, params, filename):
        """Stream `_PRICEFILE_SQL` as CSV via Postgres COPY (optionally
        gzipped). Shared by the per-brand and per-family download routes."""
        # Only compress when the client actually advertises support for it
        # (a plain curl request sends no Accept-Encoding and must get a
        # plain .csv, not a gzip blob saved with a .csv extension). Odoo's
        # HTTPRequest wrapper only proxies a fixed attribute whitelist that
        # excludes accept_encodings, so read the raw header instead.
        gz_ok = 'gzip' in request.httprequest.headers.get('Accept-Encoding', '')

        # Raw cursor bypasses the ORM's pending-write buffer: flush so stored
        # computed fields (baf_sales_column_key, baf_brand_family) and any
        # pending write() are visible to the COPY below.
        request.env.flush_all()

        buf = io.BytesIO()
        raw_cur = request.env.cr._cnx.cursor()
        try:
            # JIT costs ~0.9s on this query and buys nothing. The raw cursor
            # shares the ORM's connection, so SET LOCAL stays in effect for
            # the rest of this request's transaction, not just this cursor.
            raw_cur.execute("SET LOCAL jit = off")
            select_sql = raw_cur.mogrify(sql, params).decode()
            copy_sql = "COPY (%s) TO STDOUT WITH (FORMAT CSV, HEADER true)" % select_sql
            if gz_ok:
                with gzip.GzipFile(fileobj=buf, mode='wb') as gz:
                    gz.write(b"\xef\xbb\xbf")
                    raw_cur.copy_expert(copy_sql, gz)
            else:
                buf.write(b"\xef\xbb\xbf")
                raw_cur.copy_expert(copy_sql, buf)
        finally:
            raw_cur.close()

        data = buf.getvalue()
        buf.close()

        headers = [
            ('Content-Type', 'text/csv; charset=utf-8'),
            ('Content-Length', str(len(data))),
            ('Content-Disposition', content_disposition(filename)),
        ]
        if gz_ok:
            headers.append(('Content-Encoding', 'gzip'))
        return request.make_response(data, headers=headers)
