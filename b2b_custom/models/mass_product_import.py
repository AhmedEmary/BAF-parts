import base64
import io
import csv
import json
import logging
import time
from psycopg2.extras import execute_values
from odoo import models, fields, api, _
from odoo.exceptions import UserError
from odoo.tools import config
from odoo.tools.sql import column_exists
from odoo.addons.general_system_custom.models.baf_product_pricing import resolve_baf_brand_info

try:
    import openpyxl
except ImportError:
    openpyxl = None

_logger = logging.getLogger(__name__)

# Values that mean "missing" in the source spreadsheet. Treated as if the
# cell were empty (e.g. German "#NV" = nicht verfügbar / not available).
SENTINEL_VALUES = {
    '#NV', '#N/V', '#N/A', 'N/A', 'NA', 'NULL', 'NONE', 'NIL',
    '-', '--', '?', 'TBD', 'TBA', 'NV',
}


def _is_sentinel(value):
    return bool(value) and str(value).strip().upper() in SENTINEL_VALUES


FUZZY_MAP = {
    'sku':           ['sku', 'internal reference', 'default_code', 'part number', 'oem'],
    'brand':         ['brand', 'make', 'manufacturer'],
    'name':          ['name', 'product name', 'description', 'title'],
    'price':         ['price', 'sales price', 'list_price', 'retail', 'upe'],
    'uos':           ['unit of sales', 'unit of sale', 'uos', 'moq', 'min qty'],
    'baf_disc_code': ['discount code', 'disc code', 'discount_code', 'baf_discount_code', 'rabattcode'],
    'baf_type_code': ['type code', 'type_code', 'baf_type_code', 'type'],
    'baf_mod':       ['mod', 'baf_mod', 'motorrad', 'motorcycle'],
    'supplier_route':['supplier route', 'supplier_route', 'route'],
    'origin':        ['origin', 'origine', 'country', 'country_code', 'origin_country', 'coo'],
    'hs_code':       ['hs code', 'hscode', 'hs_code', 'tariff'],
    'surcharge':     ['surcharge', 'fee', 'core charge'],
    'weight':        ['weight', 'kg', 'mass'],
    'height':        ['height', 'hoehe', 'höhe', 'h cm', 'height cm'],
    'width':         ['width', 'breite', 'w cm', 'width cm'],
    'length':        ['length', 'laenge', 'länge', 'l cm', 'length cm', 'depth', 'tiefe'],
    'replaced_by':   ['replaced by', 'replaced_by', 'replacement', 'superseded by', 'nachfolger', 'ersetzt durch'],
    'is_storable':   ['storable', 'is_storable', 'track inventory', 'inventory tracking'],
    'is_published':  ['published', 'is_published', 'website published', 'publish'],
}


class MassProductImport(models.TransientModel):
    _name = 'mass.product.import'
    _description = 'Mass Product Import via Direct SQL'

    state = fields.Selection([
        ('upload', 'Upload File'),
        ('mapping', 'Map Columns')
    ], string='Status', default='upload')

    file_data = fields.Binary('Excel/CSV File', required=True)
    file_name = fields.Char('File Name')
    mapping_ids = fields.One2many('mass.product.import.mapping', 'import_id', string='Column Mappings')

    brand_source = fields.Selection([
        ('excel', 'Read Brand from File'),
        ('manual', 'Select Brand Manually (applies to all rows)'),
    ], string='Brand Source', default='excel', required=True)

    manual_brand_id = fields.Many2one(
        'product.brand',
        string='Brand',
        help="This brand will be applied to every product in the file.",
    )

    storable_source = fields.Selection([
        ('excel', 'Read from File'),
        ('manual', 'Set Manually (applies to all rows)'),
    ], string='Track Inventory Source', default='manual', required=True)

    manual_is_storable = fields.Boolean(
        string='Track Inventory',
        default=True,
        help="When set, products are storable (inventory tracked). When unset, products are consumable.",
    )

    published_source = fields.Selection([
        ('excel', 'Read from File'),
        ('manual', 'Set Manually (applies to all rows)'),
    ], string='Published Status Source', default='manual', required=True)

    manual_published = fields.Boolean(
        string='Is Published',
        default=False,
        help="When set, products will be published on the website.",
    )

    manual_tax_ids = fields.Many2many(
        'account.tax',
        'mass_product_import_tax_rel',
        'import_id',
        'tax_id',
        string='Customer Taxes',
        domain="[('type_tax_use', '=', 'sale')]",
        help="When set, these taxes are applied to every product imported "
             "(existing customer taxes on matched products are replaced).",
    )

    def _guess_mapped_field(self, header):
        clean_header = str(header).lower().strip()
        for field_key, fuzzy_list in FUZZY_MAP.items():
            if clean_header in fuzzy_list:
                return field_key
        return False

    def _build_mapping_lines(self, headers):
        mapping_lines = []
        for index, header in enumerate(headers):
            guessed_field = self._guess_mapped_field(header)
            # In manual mode, don't auto-map the column that is being
            # overridden by a single manually-chosen value.
            if guessed_field == 'brand' and self.brand_source == 'manual':
                guessed_field = False
            if guessed_field == 'is_storable' and self.storable_source == 'manual':
                guessed_field = False
            if guessed_field == 'is_published' and self.published_source == 'manual':
                guessed_field = False
            mapping_lines.append((0, 0, {
                'column_index': index,
                'file_column_name': str(header),
                'field_name': guessed_field,
            }))

        return mapping_lines

    def _decode_csv_text(self, binary_data):
        try:
            return base64.b64decode(binary_data).decode('utf-8-sig')
        except UnicodeDecodeError:
            return base64.b64decode(binary_data).decode('ISO-8859-1')

    def _read_import_source(self, binary_data, file_name):
        file_name_lower = (file_name or '').lower()
        if file_name_lower.endswith('.csv'):
            file_content = self._decode_csv_text(binary_data)
            csv_reader = csv.reader(io.StringIO(file_content))
            headers = next(csv_reader)
            total_rows = max(0, file_content.count('\n') - 1)
            return headers, csv_reader, total_rows

        file_content = io.BytesIO(base64.b64decode(binary_data))
        wb = openpyxl.load_workbook(filename=file_content, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter)
        total_rows = None
        try:
            if ws.max_row:
                total_rows = max(0, ws.max_row - 1)
        except Exception:
            total_rows = None
        return headers, rows_iter, total_rows

    def _prepare_import_context(self):
        uom_record = self.env.ref('uom.product_uom_unit', raise_if_not_found=False)
        uom_id = uom_record.id if uom_record else self.env['uom.uom'].search([], limit=1).id

        categ_record = self.env['product.category'].search([('name', '=ilike', 'Goods')], limit=1)
        if not categ_record:
            categ_record = self.env['product.category'].create({'name': 'Goods'})
        categ_id = categ_record.id

        if not uom_id or not categ_id:
            raise UserError("Missing Category or Unit of Measure in the database!")

        brand_cache = {
            b.name.strip().lower(): b.id
            for b in self.env['product.brand'].search([])
            if b.name
        }
        country_cache = {
            c.code.upper(): c.id
            for c in self.env['res.country'].search([])
            if c.code
        }
        replacement_cache = {}

        # is_published physically belongs to product_template (website_sale).
        has_is_published = column_exists(self.env.cr, 'product_template', 'is_published')

        return {
            'uom_id': uom_id,
            'categ_id': categ_id,
            'brand_cache': brand_cache,
            'country_cache': country_cache,
            'replacement_cache': replacement_cache,
            'has_is_published': has_is_published,
            'manual_brand_id': self.manual_brand_id.id if self.brand_source == 'manual' else None,
            'manual_brand_name': self.manual_brand_id.name if self.brand_source == 'manual' else None,
            'manual_tax_ids': self.manual_tax_ids.ids,
        }

    def _get_column_map(self):
        return {line.field_name: line.column_index for line in self.mapping_ids if line.field_name}

    def _get_column_indices(self, col_map):
        return {
            'sku_idx': col_map.get('sku'),
            'brand_idx': col_map.get('brand'),
            'name_idx': col_map.get('name'),
            'price_idx': col_map.get('price'),
            'uos_idx': col_map.get('uos'),
            'baf_disc_idx': col_map.get('baf_disc_code'),
            'baf_type_idx': col_map.get('baf_type_code'),
            'baf_mod_idx': col_map.get('baf_mod'),
            'route_idx': col_map.get('supplier_route'),
            'origin_idx': col_map.get('origin'),
            'hs_code_idx': col_map.get('hs_code'),
            'surcharge_idx': col_map.get('surcharge'),
            'weight_idx': col_map.get('weight'),
            'height_idx': col_map.get('height'),
            'width_idx': col_map.get('width'),
            'length_idx': col_map.get('length'),
            'replaced_by_idx': col_map.get('replaced_by'),
            'is_storable_idx': col_map.get('is_storable'),
            'is_published_idx': col_map.get('is_published'),
        }

    def _get_cell_value(self, row_data, idx, default=''):
        if idx is not None and len(row_data) > idx and row_data[idx] not in (None, ''):
            return str(row_data[idx]).strip()
        return default

    def _parse_float(self, value):
        if not value:
            return None
        try:
            return float(str(value).replace(',', ''))
        except ValueError:
            return None

    def _parse_int(self, value):
        if not value:
            return None
        try:
            return int(float(str(value).replace(',', '')))
        except ValueError:
            return None

    def _normalize_bool(self, value):
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        v = str(value).strip().lower()
        if v in ('1', 'true', 'yes', 'y', 'oui', 'si'):
            return True
        if v in ('0', 'false', 'no', 'n', 'non'):
            return False
        return None

    def _normalize_baf_mod(self, baf_mod):
        baf_mod = str(baf_mod or 'car').lower()
        if baf_mod in ('motorrad', 'motorcycle', 'moto'):
            return 'motorcycle'
        if baf_mod == 'sb':
            return 'sb'
        return 'car'

    def _normalize_supplier_route(self, route):
        route = str(route or 'de_table').lower()
        return route if route in ('de_table', 'eu_direct') else 'de_table'

    def _resolve_brand_id(self, brand_name, brand_cache):
        brand_key = brand_name.lower()
        if brand_key not in brand_cache:
            new_brand = self.env['product.brand'].create({'name': brand_name})
            brand_cache[brand_key] = new_brand.id
        return brand_cache[brand_key]

    def _resolve_origin_id(self, origin_input, country_cache):
        if not origin_input:
            return None
        country_code = origin_input.strip().upper()
        if country_code in country_cache:
            return country_cache[country_code]

        target_country = self.env['res.country'].search([('code', '=', country_code)], limit=1)
        if target_country:
            country_cache[country_code] = target_country.id
            return target_country.id
        return None

    def _build_row_payload(self, row, indices, ctx):
        sku = self._get_cell_value(row, indices['sku_idx'])
        if not sku:
            return None
        if _is_sentinel(sku):
            return 'sentinel'

        # Brand comes either from the file or from a single manually-chosen brand.
        if self.brand_source == 'manual':
            brand_id = ctx['manual_brand_id']
            brand_name = ctx['manual_brand_name']
        else:
            brand_name = self._get_cell_value(row, indices['brand_idx'])
            if not brand_name:
                return None
            if _is_sentinel(brand_name):
                return 'sentinel'
            brand_id = self._resolve_brand_id(brand_name, ctx['brand_cache'])
        origin_id = self._resolve_origin_id(
            self._get_cell_value(row, indices['origin_idx']),
            ctx['country_cache'],
        )

        raw_name = self._get_cell_value(row, indices['name_idx'])
        if not raw_name or _is_sentinel(raw_name):
            raw_name = f"{brand_name} {sku}"

        price = self._parse_float(self._get_cell_value(row, indices['price_idx']))
        uos = self._parse_int(self._get_cell_value(row, indices['uos_idx']))
        surcharge = self._parse_float(self._get_cell_value(row, indices['surcharge_idx'])) or 0.0
        weight = self._parse_float(self._get_cell_value(row, indices['weight_idx'])) or 0.0
        height = self._parse_float(self._get_cell_value(row, indices['height_idx'])) or 0.0
        width = self._parse_float(self._get_cell_value(row, indices['width_idx'])) or 0.0
        length = self._parse_float(self._get_cell_value(row, indices['length_idx'])) or 0.0
        # h/w/l in cm → volume in cm3
        volume = height * width * length
        hs_code = self._get_cell_value(row, indices['hs_code_idx'], None)

        replaced_by_sku_raw = self._get_cell_value(row, indices['replaced_by_idx'])
        replaced_by_sku = (
            replaced_by_sku_raw.strip()
            if replaced_by_sku_raw and not _is_sentinel(replaced_by_sku_raw)
            else None
        )
        replaced_by_id = None
        replacement_created = False
        if replaced_by_sku:
            replaced_by_id, replacement_created = self._ensure_replacement_template(
                replaced_by_sku,
                brand_id,
                brand_name,
                ctx['uom_id'],
                ctx['categ_id'],
                ctx['replacement_cache'],
            )

        baf_disc = self._get_cell_value(row, indices['baf_disc_idx'], '0') or '0'
        baf_type = self._parse_int(self._get_cell_value(row, indices['baf_type_idx'])) or 0
        baf_mod = self._normalize_baf_mod(self._get_cell_value(row, indices['baf_mod_idx'], 'car'))
        route = self._normalize_supplier_route(self._get_cell_value(row, indices['route_idx'], 'de_table'))
        default_code = self._compute_default_code(brand_name, sku)
        computed_col_key, computed_family = resolve_baf_brand_info(brand_name, baf_type, baf_mod)

        # Track Inventory: from a manual override or read from the file.
        if self.storable_source == 'manual':
            is_storable = self.manual_is_storable
        else:
            is_storable = self._normalize_bool(
                self._get_cell_value(row, indices['is_storable_idx'], None)
            )
            if is_storable is None:
                is_storable = True

        # Published status: from a manual override or read from the file.
        if self.published_source == 'manual':
            is_published = self.manual_published
        else:
            is_published = self._normalize_bool(
                self._get_cell_value(row, indices['is_published_idx'], None)
            )
            if is_published is None:
                is_published = False

        return {
            'name_json': json.dumps({"en_US": raw_name}),
            'default_code': default_code,
            'sku': sku,
            'brand_name': brand_name,
            'brand_id': brand_id,
            'price': price,
            'uos': uos,
            'origin_id': origin_id,
            'hs_code': hs_code,
            'surcharge': surcharge,
            'weight': weight,
            'height': height,
            'width': width,
            'length': length,
            'volume': volume,
            'replaced_by_id': replaced_by_id,
            'replaced_by_present': bool(replaced_by_sku),
            'replacement_created': replacement_created,
            'baf_disc': baf_disc,
            'baf_type': baf_type,
            'baf_mod': baf_mod,
            'route': route,
            'computed_col_key': computed_col_key,
            'computed_family': computed_family,
            'is_storable': is_storable,
            'is_published': is_published,
        }

    def _build_template_upsert_tuple(self, payload, ctx):
        template_tuple = (
            payload['name_json'],      # 1. name
            payload['default_code'],   # 2. default_code
            payload['sku'],            # 3. sku
            payload['brand_id'],       # 4. brand
            payload['price'],          # 5. list_price
            'consu',                   # 6. type
            payload['is_storable'],    # 7. is_storable
            ctx['uom_id'],             # 8. uom_id
            ctx['categ_id'],           # 9. categ_id
            True,                      # 10. active
            'no',                      # 11. service_tracking
            'none',                    # 12. tracking
            0.0,                       # 13. base_unit_count
            True,                      # 14. sale_ok
            True,                      # 15. purchase_ok
            payload['baf_disc'],       # 16. baf_discount_code
            payload['baf_type'],       # 17. baf_type_code
            payload['baf_mod'],        # 18. baf_mod
            payload['route'],          # 19. supplier_route
            payload['computed_col_key'],   # 20. baf_column_key
            payload['computed_family'],    # 21. baf_brand_family
            payload['origin_id'],      # 22. origin
            payload['hs_code'],        # 23. hs_code
            payload['surcharge'],      # 24. surcharge
            payload['weight'],         # 25. weight
            payload['height'],         # 26. height
            payload['width'],          # 27. width
            payload['length'],         # 28. length
            payload['volume'],         # 29. volume
            payload['replaced_by_id'], # 30. replaced_by_id
            'order',                   # 31. invoice_policy
            fields.Datetime.now(),     # 32. publish_date
        )
        if ctx['has_is_published']:
            template_tuple += (payload['is_published'],)  # 33. is_published
        return template_tuple

    def _get_product_template_upsert_query(self, has_is_published=False):
        opt_insert = ', is_published' if has_is_published else ''
        opt_update = (
            ',\n                        is_published     = EXCLUDED.is_published'
            if has_is_published else ''
        )
        return f"""
                    INSERT INTO product_template (
                        name, default_code, sku, brand, list_price,
                        type, is_storable, uom_id, categ_id, active, service_tracking,
                        tracking, base_unit_count,
                        sale_ok, purchase_ok,
                        baf_discount_code, baf_type_code, baf_mod, supplier_route,
                        baf_column_key, baf_brand_family,
                        origin, hs_code, surcharge, weight,
                        height, width, length, volume,
                        replaced_by_id,
                        invoice_policy, publish_date{opt_insert}
                    ) VALUES %s
                    ON CONFLICT (default_code) DO UPDATE SET
                        name             = EXCLUDED.name,
                        list_price       = EXCLUDED.list_price,
                        sku              = EXCLUDED.sku,
                        brand            = EXCLUDED.brand,
                        is_storable      = EXCLUDED.is_storable,
                        baf_discount_code= EXCLUDED.baf_discount_code,
                        baf_type_code    = EXCLUDED.baf_type_code,
                        baf_mod          = EXCLUDED.baf_mod,
                        supplier_route   = EXCLUDED.supplier_route,
                        baf_column_key   = EXCLUDED.baf_column_key,
                        baf_brand_family = EXCLUDED.baf_brand_family,
                        origin           = EXCLUDED.origin,
                        hs_code          = EXCLUDED.hs_code,
                        surcharge        = EXCLUDED.surcharge,
                        weight           = EXCLUDED.weight,
                        height           = EXCLUDED.height,
                        width            = EXCLUDED.width,
                        length           = EXCLUDED.length,
                        volume           = EXCLUDED.volume,
                        replaced_by_id   = COALESCE(EXCLUDED.replaced_by_id, product_template.replaced_by_id),
                        invoice_policy   = EXCLUDED.invoice_policy,
                        publish_date     = EXCLUDED.publish_date{opt_update}
                    RETURNING id, default_code;
                """

    def action_read_headers(self):
        self.ensure_one()
        if not self.file_name or not self.file_name.lower().endswith(('.xlsx', '.csv')):
            raise UserError(_("Unsupported file format. Please upload a .csv or .xlsx file."))

        if self.brand_source == 'manual' and not self.manual_brand_id:
            raise UserError(_("Please select a brand before reading the file."))

        headers = self._get_file_headers()
        mapping_lines = self._build_mapping_lines(headers)

        self.mapping_ids = False
        self.mapping_ids = mapping_lines
        self.state = 'mapping'
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'mass.product.import',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_reset(self):
        self.state = 'upload'
        self.mapping_ids = False
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'mass.product.import',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_import_direct(self):
        self.ensure_one()
        mapped_fields = [line.field_name for line in self.mapping_ids if line.field_name]
        if 'sku' not in mapped_fields:
            raise UserError(_("You must map the 'SKU' column to proceed!"))
        if self.brand_source == 'excel':
            if 'brand' not in mapped_fields:
                raise UserError(_(
                    "You must map a 'Brand' column when using 'Read Brand from File'.\n"
                    "Alternatively, switch to 'Select Brand Manually' on the previous step."
                ))
        elif self.brand_source == 'manual' and not self.manual_brand_id:
            raise UserError(_("Please select a Brand (go back and choose one)."))

        attachment = self.env['ir.attachment'].create({
            'name': self.file_name,
            'type': 'binary',
            'datas': self.file_data,
            'res_model': 'mass.product.import',
            'res_id': self.id,
        })
        self._process_direct_sql(attachment.id)
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Import Completed',
                'message': 'Your products were successfully imported!',
                'type': 'success',
                'sticky': False,
            }
        }

    def _get_file_headers(self):
        headers, _rows, _total_rows = self._read_import_source(self.file_data, self.file_name)
        return headers

    def _compute_default_code(self, brand_name, sku):
        if not sku or not brand_name:
            return None
        brand_name = str(brand_name).strip()
        sku = str(sku).strip()
        prefix = brand_name[:3].upper() if len(brand_name) >= 3 else brand_name.upper()
        return f"{prefix}_{sku}"

    def _compute_barcode_from_code(self, default_code):
        if not default_code or '_' not in default_code:
            return None
        number_part = default_code.split('_', 1)[-1]
        prefix = default_code[:3].upper()
        return number_part.zfill(9) if prefix in ['MAS', 'FER'] else number_part

    def _compute_raw_column_key(self, brand_name, type_code, mod):
        """Return only the column_key (kept for backward compat)."""
        column_key, _family = resolve_baf_brand_info(brand_name, type_code, mod)
        return column_key

    def _push_import_progress(self, done, total, started_at):
        """Send a toast with current progress over the user's bus channel.

        Bus messages are flushed on commit, so calling this just before
        cr.commit() in _execute_sql_batch delivers the toast immediately.
        """
        elapsed = time.time() - started_at
        if total and total > 0:
            pct = min(100.0, (done / total) * 100.0)
            eta_s = int((elapsed / done) * (total - done)) if done else 0
            message = _(
                "Imported %(d)s / %(t)s rows (%(p).1f%%). ETA ~%(em)dm %(es)ds"
            ) % {
                'd': f'{done:,}', 't': f'{total:,}',
                'p': pct, 'em': eta_s // 60, 'es': eta_s % 60,
            }
        else:
            message = _("Imported %(d)s rows so far (%(e).0fs elapsed)") % {
                'd': f'{done:,}', 'e': elapsed,
            }

        self.env['bus.bus']._sendone(
            self.env.user.partner_id,
            'simple_notification',
            {
                'title': _("Mass Product Import"),
                'message': message,
                'type': 'info',
                'sticky': False,
            },
        )

    def _ensure_replacement_template(self, replacement_sku, brand_id, brand_name, uom_id, categ_id, replacement_cache):
        """Ensure the replacement template exists *before* importing the source row.

        The replacement is always created/located using the same brand context
        as the source row, because the spreadsheet only provides the SKU. That
        guarantees a stable same-brand default_code (e.g. BOS_NEW001) and lets
        a later real import update the placeholder in place.

        Returns: (replacement_template_id, created_now)
        """
        cache_key = (brand_id, replacement_sku)
        if cache_key in replacement_cache:
            return replacement_cache[cache_key], False

        default_code = self._compute_default_code(brand_name, replacement_sku)
        cr = self.env.cr
        cr.execute(
            "SELECT id FROM product_template WHERE default_code = %s LIMIT 1",
            (default_code,),
        )
        existing = cr.fetchone()
        if existing:
            replacement_cache[cache_key] = existing[0]
            return existing[0], False

        placeholder_name = json.dumps({"en_US": f"{brand_name} {replacement_sku}"})
        column_key, family = resolve_baf_brand_info(brand_name, 0, 'car')
        insert_query = """
            INSERT INTO product_template (
                name, default_code, sku, brand, list_price,
                type, is_storable, uom_id, categ_id, active, service_tracking,
                tracking, base_unit_count,
                sale_ok, purchase_ok,
                baf_discount_code, baf_type_code, baf_mod, supplier_route,
                baf_column_key, baf_brand_family,
                origin, hs_code, surcharge, weight,
                height, width, length, volume,
                invoice_policy, publish_date
            ) VALUES %s
            ON CONFLICT (default_code) DO UPDATE SET
                name = EXCLUDED.name,
                sku = EXCLUDED.sku,
                brand = EXCLUDED.brand,
                baf_column_key = EXCLUDED.baf_column_key,
                baf_brand_family = EXCLUDED.baf_brand_family
            RETURNING id;
        """
        created = execute_values(cr._obj, insert_query, [(
            placeholder_name,
            default_code,
            replacement_sku,
            brand_id,
            0.0,
            'consu',
            True,
            uom_id,
            categ_id,
            True,
            'no',
            'none',
            0.0,
            True,
            True,
            '0',
            0,
            'car',
            'de_table',
            column_key,
            family,
            None,
            None,
            0.0,
            0.0,
            0.0,
            0.0,
            0.0,
            0.0,
            'order',
            fields.Datetime.now(),
        )], fetch=True)
        replacement_id = created[0][0]
        cr.execute("""
            INSERT INTO product_product (product_tmpl_id, default_code, active, base_unit_count, weight)
            SELECT pt.id, pt.default_code, true, 0.0, pt.weight
            FROM product_template pt
            LEFT JOIN product_product pp ON pp.product_tmpl_id = pt.id
            WHERE pt.id = %s AND pp.id IS NULL;
        """, (replacement_id,))
        replacement_cache[cache_key] = replacement_id
        return replacement_id, True

    def _process_direct_sql(self, attachment_id):
        total_start_time = time.time()
        attachment = self.env['ir.attachment'].browse(attachment_id)
        if not attachment:
            return

        ctx = self._prepare_import_context()
        _headers, rows, total_rows = self._read_import_source(attachment.datas, self.file_name)
        col_map = self._get_column_map()
        indices = self._get_column_indices(col_map)
        upsert_query = self._get_product_template_upsert_query(ctx['has_is_published'])

        batch_size = 50000
        data_batch_dict = {}
        batch_counter = 0
        rows_committed = 0
        last_pct_pushed = 0.0
        replacement_links_set = 0
        replacement_placeholders_created = 0

        # Kick-off toast so the user sees something within seconds.
        self._push_import_progress(0, total_rows, total_start_time)

        skipped_sentinel = 0
        for row in rows:
            if not row:
                continue

            payload = self._build_row_payload(row, indices, ctx)
            if payload == 'sentinel':
                skipped_sentinel += 1
                continue
            if not payload:
                continue

            if payload['replaced_by_present']:
                replacement_links_set += 1
            if payload['replacement_created']:
                replacement_placeholders_created += 1

            data_batch_dict[payload['default_code']] = self._build_template_upsert_tuple(payload, ctx)

            if len(data_batch_dict) >= batch_size:
                batch_counter += 1
                rows_committed += len(data_batch_dict)
                self._execute_sql_batch(upsert_query, list(data_batch_dict.values()), batch_counter, ctx['manual_tax_ids'])
                data_batch_dict.clear()
                # Throttle: push a toast only when progress crosses each 5%
                # threshold (or every 10 batches if total is unknown).
                if total_rows:
                    pct_now = (rows_committed / total_rows) * 100.0
                    if pct_now - last_pct_pushed >= 5.0:
                        self._push_import_progress(rows_committed, total_rows, total_start_time)
                        last_pct_pushed = pct_now
                elif batch_counter % 10 == 0:
                    self._push_import_progress(rows_committed, total_rows, total_start_time)

        if data_batch_dict:
            batch_counter += 1
            rows_committed += len(data_batch_dict)
            self._execute_sql_batch(upsert_query, list(data_batch_dict.values()), batch_counter)

        # Final 100% toast
        self._push_import_progress(rows_committed, total_rows or rows_committed, total_start_time)

        attachment.unlink()
        _logger.info(
            "MASS SQL IMPORT: finished in %.2f minutes (%d rows committed, "
            "%d sentinel rows skipped, %d replaced_by links processed, %d replacement placeholders created).",
            (time.time() - total_start_time) / 60.0,
            rows_committed, skipped_sentinel, replacement_links_set, replacement_placeholders_created,
        )

    def _execute_sql_batch(self, query, data_batch, batch_counter, tax_ids=None):
        self.env.flush_all()
        cr = self.env.cr
        batch_start = time.time()
        upserted_templates = execute_values(cr._obj, query, data_batch, fetch=True)
        if upserted_templates:
            template_ids = tuple([row[0] for row in upserted_templates])

            cr.execute("""
                INSERT INTO product_product (product_tmpl_id, default_code, active, base_unit_count, weight)
                SELECT pt.id, pt.default_code, true, 0.0, pt.weight
                FROM product_template pt
                LEFT JOIN product_product pp ON pp.product_tmpl_id = pt.id
                WHERE pt.id IN %s AND pp.id IS NULL;
            """, (template_ids,))

            barcode_batch = [
                (self._compute_barcode_from_code(dc), dc)
                for _, dc in upserted_templates
                if self._compute_barcode_from_code(dc)
            ]
            if barcode_batch:
                execute_values(cr._obj, """
                    UPDATE product_product AS pp
                    SET barcode = v.barcode
                    FROM (VALUES %s) AS v(barcode, default_code)
                    WHERE pp.default_code = v.default_code;
                """, barcode_batch)

            if tax_ids:
                # Replace existing customer taxes for every upserted template
                # with the user-selected set, in a single round-trip.
                cr.execute(
                    "DELETE FROM product_taxes_rel WHERE prod_id IN %s;",
                    (template_ids,),
                )
                execute_values(cr._obj, """
                    INSERT INTO product_taxes_rel (prod_id, tax_id)
                    VALUES %s
                    ON CONFLICT DO NOTHING;
                """, [(tid, tax_id) for tid in template_ids for tax_id in tax_ids])

        # Commit per batch so locks (ir_attachment, product_template) are
        # released, the auto-vacuum cron can run, and a partial import
        # survives if the request is interrupted. Skipped under tests
        # because TestCursor forbids commit/rollback.
        if not config['test_enable']:
            cr.commit()
        _logger.info(
            "MASS SQL IMPORT: batch %d (%d rows) committed in %.2fs",
            batch_counter, len(data_batch), time.time() - batch_start,
        )


class MassProductImportMapping(models.TransientModel):
    _name = 'mass.product.import.mapping'
    _description = 'Mass Product Import Mapping Line'

    import_id = fields.Many2one('mass.product.import', ondelete='cascade')
    column_index = fields.Integer("Column Position")
    file_column_name = fields.Char("Excel/CSV Header")

    field_name = fields.Selection([
        ('sku',           'SKU (Required)'),
        ('brand',         'Brand (Required)'),
        ('name',          'Product Name'),
        ('price',         'Sales Price / UPE'),
        ('uos',           'Unit of Sales'),
        ('weight',        'Weight'),
        ('height',        'Height (cm)'),
        ('width',         'Width (cm)'),
        ('length',        'Length (cm)'),
        ('surcharge',     'Surcharge'),
        ('hs_code',       'HS Code'),
        ('baf_disc_code', 'BAF Discount Code #'),
        ('baf_type_code', 'BAF Type Code (1-9)'),
        ('baf_mod',       'BAF Mod (car / motorcycle / sb)'),
        ('supplier_route','Supplier Route (de_table / eu_direct)'),
        ('origin',        'Origin (Country Code)'),
        ('replaced_by',   'Replaced By (SKU of replacement product)'),
        ('is_storable',   'Track Inventory'),
        ('is_published',  'Is Published'),
    ], string="Odoo Field")
