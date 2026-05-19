import base64
import io
import csv
import logging
import time
from psycopg2.extras import execute_values
from odoo import models, fields, _
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None

_logger = logging.getLogger(__name__)

# Auto-guess the mapped field based on the file headers.
VENDOR_FUZZY_MAP = {
    'sku':                ['sku', 'internal reference', 'default_code', 'part number', 'reference', 'oem'],
    'brand':              ['brand', 'make', 'manufacturer'],
    'price':              ['price', 'cost', 'vendor price', 'supplier price', 'purchase price', 'cost price'],
    'min_qty':            ['min qty', 'minimum quantity', 'min_qty', 'moq'],
    'delivery_lead_time': ['lead time', 'delivery lead time', 'delay', 'delivery_lead_time'],
}


class MassVendorPriceImport(models.TransientModel):
    _name = 'mass.vendor.price.import'
    _description = 'Mass Vendor Price Import'

    state = fields.Selection([
        ('upload', 'Upload File'),
        ('mapping', 'Map Columns')
    ], string='Status', default='upload')

    file_data = fields.Binary('Excel/CSV File', required=True)
    file_name = fields.Char('File Name')

    vendor_id = fields.Many2one(
        'res.partner',
        string='Vendor',
        required=True,
        domain=[('supplier_rank', '>', 0)],
        help="Select the vendor/supplier whose prices you are importing.",
    )

    mapping_ids = fields.One2many(
        'mass.vendor.price.import.mapping', 'import_id', string='Column Mappings'
    )

    def action_read_headers(self):
        """Step 1: Read file headers and prepare the mapping grid."""
        self.ensure_one()

        if not self.file_name or not self.file_name.lower().endswith(('.xlsx', '.csv')):
            raise UserError(_("Unsupported file format. Please upload a .csv or .xlsx file."))

        if not self.vendor_id:
            raise UserError(_("Please select a Vendor before reading the file."))

        headers = self._get_file_headers()

        mapping_lines = []
        for index, header in enumerate(headers):
            clean_header = str(header).lower().strip()

            guessed_field = False
            for field_key, fuzzy_list in VENDOR_FUZZY_MAP.items():
                if clean_header in fuzzy_list:
                    guessed_field = field_key
                    break

            mapping_lines.append((0, 0, {
                'column_index': index,
                'file_column_name': str(header),
                'field_name': guessed_field,
            }))

        self.mapping_ids = False
        self.mapping_ids = mapping_lines
        self.state = 'mapping'

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'mass.vendor.price.import',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_reset(self):
        """Go back to the upload step."""
        self.state = 'upload'
        self.mapping_ids = False
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'mass.vendor.price.import',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_import(self):
        """Step 2: Process the file and create/update product.supplierinfo records."""
        self.ensure_one()

        mapped_fields = [line.field_name for line in self.mapping_ids if line.field_name]

        if 'sku' not in mapped_fields:
            raise UserError(_("You must map the 'SKU' column to proceed!"))
        if 'brand' not in mapped_fields:
            raise UserError(_("You must map the 'Brand' column to proceed!"))
        if 'price' not in mapped_fields:
            raise UserError(_("You must map the 'Price' column to proceed!"))

        self._process_vendor_prices()

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Vendor Price Import Completed'),
                'message': _('Vendor prices were successfully imported/updated!'),
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }

    def _get_file_headers(self):
        file_name_lower = self.file_name.lower()
        if file_name_lower.endswith('.csv'):
            try:
                file_content = base64.b64decode(self.file_data).decode('utf-8-sig')
            except UnicodeDecodeError:
                file_content = base64.b64decode(self.file_data).decode('ISO-8859-1')
            csv_reader = csv.reader(io.StringIO(file_content))
            return next(csv_reader)
        else:
            file_content = io.BytesIO(base64.b64decode(self.file_data))
            wb = openpyxl.load_workbook(filename=file_content, read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            return next(rows_iter)

    def _compute_default_code(self, brand_name, sku):
        """Build the product code the same way the mass product import does:
        a 3-letter brand prefix joined to the SKU (e.g. BOS_12345)."""
        if not sku or not brand_name:
            return None
        brand_name = str(brand_name).strip()
        sku = str(sku).strip()
        prefix = brand_name[:3].upper() if len(brand_name) >= 3 else brand_name.upper()
        return f"{prefix}_{sku}"

    def _process_vendor_prices(self):
        total_start_time = time.time()

        # Build the column map from the user mapping.
        col_map = {line.field_name: line.column_index for line in self.mapping_ids if line.field_name}

        sku_idx = col_map.get('sku')
        brand_idx = col_map.get('brand')
        price_idx = col_map.get('price')
        min_qty_idx = col_map.get('min_qty')
        delay_idx = col_map.get('delivery_lead_time')

        # Read file rows.
        file_name_lower = self.file_name.lower()
        if file_name_lower.endswith('.csv'):
            try:
                file_content = base64.b64decode(self.file_data).decode('utf-8-sig')
            except UnicodeDecodeError:
                file_content = base64.b64decode(self.file_data).decode('ISO-8859-1')
            csv_reader = csv.reader(io.StringIO(file_content))
            next(csv_reader)  # Skip header
            all_rows = list(csv_reader)
        else:
            file_content = io.BytesIO(base64.b64decode(self.file_data))
            wb = openpyxl.load_workbook(filename=file_content, read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            next(rows_iter)  # Skip header
            all_rows = list(rows_iter)

        vendor_id = self.vendor_id.id
        default_currency_id = self.env.company.currency_id.id
        uom = self.env.ref('uom.product_uom_unit', raise_if_not_found=False)
        default_uom_id = uom.id if uom else self.env['uom.uom'].search([], limit=1).id

        def get_val(row_data, idx, default=''):
            if idx is not None and len(row_data) > idx and row_data[idx] not in (None, ''):
                return str(row_data[idx]).strip()
            return default

        def safe_float(val_str):
            if not val_str:
                return 0.0
            if isinstance(val_str, (int, float)):
                return float(val_str)
            try:
                return float(str(val_str).replace(',', ''))
            except (ValueError, TypeError):
                return 0.0

        def safe_int(val_str):
            if not val_str:
                return 0
            if isinstance(val_str, (int, float)):
                return int(val_str)
            try:
                return int(float(str(val_str).replace(',', '')))
            except (ValueError, TypeError):
                return 0

        # Build the set of default_codes we actually need from the file.
        needed_codes = set()
        for row in all_rows:
            if not row:
                continue
            sku = get_val(row, sku_idx)
            brand_name = get_val(row, brand_idx)
            if not sku or not brand_name:
                continue
            default_code = self._compute_default_code(brand_name, sku)
            if default_code:
                needed_codes.add(default_code)

        # Build a product cache only for the needed codes (avoids MemoryError).
        product_cache = {}
        if needed_codes:
            needed_codes_list = list(needed_codes)
            chunk_size = 5000
            for i in range(0, len(needed_codes_list), chunk_size):
                chunk = needed_codes_list[i:i + chunk_size]
                self.env.cr.execute(
                    """
                    SELECT pp.id, pp.default_code, pp.product_tmpl_id
                    FROM product_product pp
                    WHERE pp.default_code IN %s
                    """,
                    (tuple(chunk),)
                )
                for pp_id, code, tmpl_id in self.env.cr.fetchall():
                    if code:
                        product_cache[code.strip()] = {
                            'product_id': pp_id,
                            'product_tmpl_id': tmpl_id,
                        }

        # Delete existing supplierinfo for this vendor + matched products via direct SQL.
        if product_cache:
            product_ids_to_clean = [v['product_id'] for v in product_cache.values()]
            if product_ids_to_clean:
                self.env.cr.execute(
                    """
                    DELETE FROM product_supplierinfo
                    WHERE partner_id = %s AND product_id IN %s
                    """,
                    (vendor_id, tuple(product_ids_to_clean))
                )

        # Process rows and build tuples for a direct SQL insert.
        sql_rows = []
        skipped_rows = []

        for row_num, row in enumerate(all_rows, start=2):
            if not row:
                continue

            sku = get_val(row, sku_idx)
            brand_name = get_val(row, brand_idx)
            price_str = get_val(row, price_idx)

            if not sku or not brand_name or not price_str:
                continue

            default_code = self._compute_default_code(brand_name, sku)
            product_data = product_cache.get(default_code)

            if not product_data:
                skipped_rows.append(row_num)
                continue

            price = safe_float(price_str)
            min_qty = safe_int(get_val(row, min_qty_idx, '0')) if min_qty_idx is not None else 0
            delay = safe_int(get_val(row, delay_idx, '1')) if delay_idx is not None else 1

            sql_rows.append((
                vendor_id,
                product_data['product_tmpl_id'],
                product_data['product_id'],
                price,
                min_qty or 0.0,
                delay or 1,
                default_currency_id,
                default_uom_id,
                1,  # sequence
            ))

        # Batch insert via direct SQL.
        if sql_rows:
            self.env.flush_all()
            cr = self.env.cr

            insert_query = """
                INSERT INTO product_supplierinfo
                    (partner_id, product_tmpl_id, product_id, price, min_qty, delay, currency_id, product_uom_id, sequence)
                VALUES %s
            """

            batch_size = 5000
            for i in range(0, len(sql_rows), batch_size):
                batch = sql_rows[i:i + batch_size]
                execute_values(cr._obj, insert_query, batch)

            self.env.invalidate_all()

        _logger.info(
            "MASS VENDOR PRICE IMPORT: finished in %.2fs (%d vendor prices imported, "
            "%d rows skipped because the product was not found).",
            time.time() - total_start_time, len(sql_rows), len(skipped_rows),
        )


class MassVendorPriceImportMapping(models.TransientModel):
    _name = 'mass.vendor.price.import.mapping'
    _description = 'Mass Vendor Price Import Mapping Line'

    import_id = fields.Many2one('mass.vendor.price.import', ondelete='cascade')
    column_index = fields.Integer("Column Position")
    file_column_name = fields.Char("Excel/CSV Header")

    field_name = fields.Selection([
        ('sku',                'SKU (Required)'),
        ('brand',              'Brand (Required)'),
        ('price',              'Vendor Price (Required)'),
        ('min_qty',            'Minimum Quantity'),
        ('delivery_lead_time', 'Delivery Lead Time (days)'),
    ], string="Odoo Field", help="Select the Odoo field this column maps to")
